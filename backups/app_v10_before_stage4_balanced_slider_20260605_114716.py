"""
app_v10.py — Stirling Engine Simulator v10 (Streamlit UI)
==========================================================
v10 changes vs v9.4:
  - Default heat input raised to 1000 W (no red deficit on load)
  - No-regen P-V overlay on Schmidt and Both tabs
  - Nonlinear Q_in sweep: With vs Without Regenerator
  - Optimization: Stage 2 (Prototype 2), Stage 3 (Full Geometry), Stage 4 (Operating Conditions)
  - Power–Efficiency map, Volume plots, Geometry Summary Table
  - Frequency & Torque Sweep across designs
  - Unified export_excel_v10 function
  - Animation gauge overlap fix
"""

import math, io, warnings, tempfile, os
from datetime import datetime

import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import streamlit as st

from animation_v10 import build_engine_animation
from physics_v10 import (
    PROTOTYPE, GASES,
    to_si, build_geometry,
    simulate, simulate_fixed_heat,
    validate_mass_conservation, validate_first_law,
    validate_carnot, validate_pressure_scaling,
    sweep_qin, frequency_torque_sweep,
    pressure_sweep, gas_comparison,
)
from optimization_v10 import (
    OPTIMIZABLE_PARAMS, geometry_sensitivity,
    coarse_fine_search, lhs_search, bayesian_search,
    prototype2_search, stage3_search, stage4_search,
    stage3_search_named, stage4_search_named,
    STAGE2_LOCKED, STAGE4_OPEN,
)


def _freeze_dict(d):
    return tuple(sorted((k, v) for k, v in d.items() if isinstance(v, (int, float, str, bool))))

warnings.filterwarnings("ignore", category=UserWarning)


# ── Helper: Power–Efficiency map ─────────────────────────────────────────────
def plot_power_efficiency_map(all_results, named_points, title="Power–Efficiency Map", Q_in_max=1500.0):
    """
    Power–Efficiency map.

    Gray dots = feasible candidates.
    Red x = infeasible candidates, Q_in > Q_in_max.
    Named points are selected only from feasible candidates.
    """
    fig, ax = plt.subplots(figsize=(9, 6))

    feasible_x, feasible_y = [], []
    infeasible_x, infeasible_y = [], []
    all_x, all_y, all_q = [], [], []

    if all_results:
        for item in all_results:
            if not isinstance(item, (tuple, list)) or len(item) < 3:
                continue

            _, _, losses = item[0], item[1], item[2]
            if not losses:
                continue

            P = losses.get('P_brake', None)
            eta = losses.get('eta_brake', None)
            q = losses.get('Q_in_W', None)

            if P is None or eta is None or q is None:
                continue

            if P <= 0:
                continue

            all_x.append(P)
            all_y.append(eta * 100)
            all_q.append(q)

            if q <= Q_in_max:
                feasible_x.append(P)
                feasible_y.append(eta * 100)
            else:
                infeasible_x.append(P)
                infeasible_y.append(eta * 100)

    if feasible_x:
        ax.scatter(feasible_x, feasible_y, s=14, color='#BDBDBD',
                   alpha=0.55, label='Feasible candidates', zorder=2)

    if infeasible_x:
        ax.scatter(infeasible_x, infeasible_y, s=22, color='#EF9A9A',
                   alpha=0.55, marker='x',
                   label='Infeasible: Q_in > budget', zorder=1)

    # Heat-budget frontier: approximate Q_in = Q_in_max boundary.
    # We bin candidates along P_brake and connect the point in each bin
    # whose Q_in is closest to Q_in_max.
    if len(all_x) >= 10:
        try:
            import numpy as _np
            xs = _np.array(all_x)
            ys = _np.array(all_y)
            qs = _np.array(all_q)

            bins = _np.linspace(xs.min(), xs.max(), 18)
            bx, by = [], []

            for i in range(len(bins) - 1):
                mask = (xs >= bins[i]) & (xs < bins[i + 1])
                if not mask.any():
                    continue

                idxs = _np.where(mask)[0]
                closest = idxs[_np.argmin(_np.abs(qs[idxs] - Q_in_max))]

                # Only use points reasonably close to the heat boundary
                if abs(qs[closest] - Q_in_max) / Q_in_max < 0.25:
                    bx.append(xs[closest])
                    by.append(ys[closest])

            if len(bx) >= 3:
                order = _np.argsort(bx)
                bx = _np.array(bx)[order]
                by = _np.array(by)[order]
                ax.plot(
                    bx, by,
                    color='#FF9800',
                    linestyle='--',
                    linewidth=2.0,
                    label=f'Heat-budget frontier: Q_in ≈ {Q_in_max:.0f} W',
                    zorder=4
                )
        except Exception:
            pass

    style_map = {
        'proto1':    ('#1565C0', 'o', 130, 'Prototype 1'),
        'proto2':    ('#2E7D32', 's', 130, 'Prototype 2'),
        'max_power': ('#C62828', '*', 220, 'Max Power'),
        'max_eta':   ('#6A1B9A', '^', 170, 'Max Efficiency'),
        'balanced':  ('#E65100', 'D', 150, 'Balanced'),
    }

    for key, losses in named_points.items():
        if losses is None:
            continue

        P = losses.get('P_brake', None)
        eta = losses.get('eta_brake', None)
        q = losses.get('Q_in_W', None)

        if P is None or eta is None or q is None or P <= 0:
            continue

        color, marker, size, label = style_map.get(key, ('#333333', 'o', 100, key))
        edge = 'black' if q <= Q_in_max else 'red'

        ax.scatter(P, eta * 100,
                   s=size, color=color, marker=marker,
                   edgecolors=edge, linewidths=1.4,
                   label=label,
                   zorder=5)

        ax.annotate(label, (P, eta * 100),
                    textcoords="offset points", xytext=(7, 7),
                    fontsize=8)

    ax.set(xlabel='P_brake [W]', ylabel='η_net [%]', title=title)
    ax.legend(fontsize=8)
    ax.grid(alpha=0.3)
    plt.tight_layout()
    return fig


# ── Helper: Geometry Summary Table ───────────────────────────────────────────
def show_geometry_summary_table(designs):
    """
    Display a multi-row geometry summary table.
    designs: dict with keys from 'proto1','proto2','max_power','max_eta','balanced'
             each value: dict with 'params', 'losses', 'geom'
    """
    label_map = {
        'proto1':    'Prototype 1',
        'proto2':    'Prototype 2',
        'max_power': 'Max Power',
        'max_eta':   'Max Efficiency',
        'balanced':  'Balanced',
    }
    rows = []
    for key, label in label_map.items():
        if key not in designs or designs[key] is None:
            continue
        d = designs[key]
        p = d['params']
        l = d['losses']
        g = d.get('geom', {})

        D_d = p.get('D_displacer', 75)   * 1e-3
        S_d = p.get('S_displacer', 101.5)* 1e-3
        D_p = p.get('D_power',     65.6) * 1e-3
        S_p = p.get('S_power',     61.6) * 1e-3
        V_swe   = math.pi*(D_d/2)**2*S_d * 1e6
        V_swc   = math.pi*(D_p/2)**2*S_p * 1e6
        V_dead  = g.get('V_r_lumped', 0) * 1e6
        alpha   = V_swc / V_swe if V_swe > 0 else 0
        q       = l['Q_in_W']
        feasible = '🟢' if q <= 1500 else '🔴'

        rows.append({
            'Design':         label,
            'V_swe [cm³]':    f"{V_swe:.1f}",
            'V_swc [cm³]':    f"{V_swc:.1f}",
            'V_total [cm³]':  f"{V_swe+V_swc:.1f}",
            'V_dead [cm³]':   f"{V_dead:.1f}",
            'α':              f"{alpha:.3f}",
            'P_brake [W]':    f"{l['P_brake']:.2f}",
            'η_net [%]':      f"{l['eta_brake']*100:.2f}",
            'Q_in [W]':       f"{q:.1f}",
            'T_h [K]':        f"{p.get('T_h', 873):.0f}",
            'Feasible':       feasible,
        })

    if rows:
        import pandas as pd
        st.dataframe(pd.DataFrame(rows).set_index('Design'))
        st.caption(
            "α = V_swc / V_swe. V_dead = total regenerator dead volume. "
            "🟢 = Q_in ≤ 1500 W. 🔴 = Q_in > 1500 W."
        )


# ── Helper: Excel export ──────────────────────────────────────────────────────
def export_excel_v10(tab_name, primary_sim, Q_in_max, extra_tables=None):
    """
    Build and return an Excel workbook as bytes.
    Generates a PV chart from primary_sim internally.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.drawing.image import Image as XLImage

    wb  = openpyxl.Workbook()
    ts  = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    L   = primary_sim['losses']
    p   = primary_sim['params']
    q   = L['Q_in_W']
    feasible = q <= Q_in_max

    hdr_fill = PatternFill("solid", fgColor="1565C0")
    hdr_font = Font(color="FFFFFF", bold=True)

    def hdr(ws, row, col, text):
        c = ws.cell(row=row, column=col, value=text)
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(horizontal='center')

    # Sheet 1: Summary
    ws1 = wb.active; ws1.title = "Summary"
    ws1.append(["Stirling Engine Simulator V10", "", ""])
    ws1.append(["Timestamp", ts]); ws1.append(["Tab", tab_name])
    ws1.append(["Model", primary_sim.get('model', '?')])
    ws1.append([])
    hdr(ws1, ws1.max_row+1, 1, "INPUT PARAMETERS")
    for key in ('D_displacer','S_displacer','D_power','S_power','phi_deg',
                'D_r','L_r','d_wire','porosity','T_h','T_k',
                'P_mean_bar','f','eps_reg','eta_mech'):
        ws1.append([key, p.get(key, ''), ''])
    ws1.append([])
    hdr(ws1, ws1.max_row+1, 1, "KEY OUTPUTS")
    ws1.append(["P_brake [W]",   round(L['P_brake'], 3)])
    ws1.append(["eta_net [%]",   round(L['eta_brake']*100, 3)])
    ws1.append(["Q_in [W]",      round(q, 2)])
    ws1.append(["M [g]",         round(L['M']*1000, 4)])
    ws1.append([])
    hdr(ws1, ws1.max_row+1, 1, "HEAT BUDGET")
    ws1.append(["Q_in_required [W]", round(q, 2)])
    ws1.append(["Q_in_max [W]",      Q_in_max])
    ws1.append(["Margin [W]",        round(Q_in_max - q, 2)])
    ws1.append(["Feasible",          "Yes" if feasible else "No"])
    ws1.append(["Note", "Q_in is net heat into the gas. "
                        "Actual heater/burner power is higher."])
    ws1.column_dimensions['A'].width = 26
    ws1.column_dimensions['B'].width = 18

    # Sheet 2: Detailed Losses
    ws2 = wb.create_sheet("Detailed Losses")
    for j, h in enumerate(["Component", "J/cycle", "W"], 1): hdr(ws2, 1, j, h)
    f_hz = p.get('f', 10)
    rows2 = [
        ("P_indicated",  L.get('W_cycle', 0),    L.get('W_cycle', 0) * f_hz),
        ("W_pump",       L.get('W_pump', 0),      L.get('W_pump', 0)  * f_hz),
        ("W_mech_loss",  L.get('W_mech_loss', 0), L.get('W_mech_loss',0)*f_hz),
        ("W_shaft",      L.get('W_shaft', 0),     L['P_brake']),
        ("Q_miss",       L.get('Q_miss', 0),      L.get('Q_miss', 0)  * f_hz),
        ("Q_cond",       L.get('Q_cond', 0),      L.get('Q_cond_W', 0)),
        ("W_leak",       L.get('W_leak', 0),      L.get('W_leak', 0)  * f_hz),
    ]
    for i, (name, jc, w) in enumerate(rows2, 2):
        ws2.cell(i, 1, name); ws2.cell(i, 2, round(jc, 5)); ws2.cell(i, 3, round(w, 3))
    ws2.column_dimensions['A'].width = 22

    # Sheet 3: Cycle Data
    ws3 = wb.create_sheet("Cycle Data")
    hdrs3 = ["theta_deg","P_bar","V_e_cm3","V_c_cm3","T_e_K","T_c_K","V_total_cm3"]
    for j, h in enumerate(hdrs3, 1): hdr(ws3, 1, j, h)
    R = primary_sim['result']
    for i in range(len(R['theta'])):
        ws3.append([
            round(math.degrees(R['theta'][i]), 2),
            round(R['P'][i] / 1e5, 5),
            round(R['V_e'][i] * 1e6, 4),
            round(R['V_c'][i] * 1e6, 4),
            round(R['T_e'][i], 2),
            round(R['T_c'][i], 2),
            round((R['V_e'][i]+R['V_c'][i]) * 1e6, 4),
        ])

    # Sheet 4: Charts — generate PV diagram from sim data
    ws4 = wb.create_sheet("Charts")
    try:
        V_tot = (R['V_e'] + R['V_c']) * 1e6
        fig_pv, ax_pv = plt.subplots(figsize=(8, 5))
        ax_pv.plot(V_tot, R['P'] / 1e5, '#1565C0', lw=2)
        ax_pv.fill(V_tot, R['P'] / 1e5, alpha=0.10, color='#1565C0')
        ax_pv.set(xlabel='V_total [cm³]', ylabel='P [bar]',
                  title=f'P-V Diagram — {tab_name}')
        ax_pv.grid(alpha=0.3)
        with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmp:
            tmp_path = tmp.name
        try:
            fig_pv.savefig(tmp_path, dpi=100, bbox_inches='tight')
            ws4.add_image(XLImage(tmp_path), 'A1')
        finally:
            plt.close(fig_pv)
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)
    except Exception:
        pass

    # Optional extra tables
    if extra_tables:
        import pandas as pd
        for sheet_name, df in extra_tables:
            ws_e = wb.create_sheet(sheet_name[:31])
            for j, col in enumerate(df.columns, 1):
                hdr(ws_e, 1, j, col)
            for i, row in enumerate(df.itertuples(index=False), 2):
                for j, val in enumerate(row, 1):
                    ws_e.cell(i, j, val)

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf.getvalue()


# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Stirling Engine Simulator v10",
    layout="wide",
    initial_sidebar_state="expanded",
)

if 'initialized' not in st.session_state:
    for k, v in PROTOTYPE.items():
        st.session_state[k] = v
    st.session_state['gap_mm']                    = PROTOTYPE['gap'] * 1e3
    st.session_state['L_displacer_effective_mm']  = PROTOTYPE['L_displacer_effective'] * 1e3
    st.session_state.initialized = True

def reset_to_prototype():
    for k, v in PROTOTYPE.items():
        st.session_state[k] = v
    st.session_state['gap_mm']                   = PROTOTYPE['gap'] * 1e3
    st.session_state['L_displacer_effective_mm'] = PROTOTYPE['L_displacer_effective'] * 1e3

col_title, col_reset = st.columns([4, 1])
with col_title:
    st.title("🔥 Stirling Engine Simulator v10")
    st.caption("Gamma-type engine — Schmidt · Adiabatic · Comparison · Optimization · Bi-directional Mode")
with col_reset:
    st.write(""); st.write("")
    if st.button("🔄 Reset to Prototype", type="secondary"):
        reset_to_prototype(); st.rerun()


# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.header("⚙️ Parameters")

    st.subheader("Geometry")
    st.number_input("Displacer diameter [mm]", 30.0, 200.0, step=1.0, key='D_displacer')
    st.number_input("Displacer stroke [mm]",   20.0, 300.0, step=1.0, key='S_displacer')
    st.number_input("Power piston diameter [mm]", 30.0, 200.0, step=1.0, key='D_power')
    st.number_input("Power piston stroke [mm]",   20.0, 300.0, step=1.0, key='S_power')
    st.number_input("Phase angle [°]", 45.0, 135.0, step=5.0, key='phi_deg')

    st.subheader("Regenerator")
    st.number_input("Regen diameter [mm]", 10.0, 150.0, step=1.0, key='D_r')
    st.number_input("Regen length [mm]",   20.0, 500.0, step=5.0, key='L_r')
    st.number_input("Wire diameter [mm]",  0.05, 5.0,   step=0.05, key='d_wire')
    st.slider("Porosity", 0.5, 0.99, step=0.01, key='porosity')

    st.subheader("Operating Conditions")
    st.selectbox("Working gas", ["Air", "Helium", "Hydrogen"], key='gas')

    driving_mode = st.radio("Driving Mode",
        ["Fixed Hot Temperature (T_h)", "Fixed Heat Input (Q_in)"],
        key='driving_mode')

    if driving_mode == "Fixed Hot Temperature (T_h)":
        st.number_input("Hot temperature T_h [K]", 400, 1500, step=10, key='T_h')
        st.slider("Max heat input budget [W]", 50, 3000, value=1000, step=10, key='Q_in_max',
            help="Heat constraint. If required heat exceeds this, available power is scaled down.")
    else:
        st.slider("Net Heat Transferred to Gas (Q_in) [W]", 50, 3000, value=1000, step=10,
                  key='Q_in_max',
                  help="Solver finds T_h that produces exactly this Q_in.")

    st.number_input("Cold temperature T_k [K]", 250, 400, step=5, key='T_k')
    st.number_input("Mean pressure [bar]", 0.5, 50.0, step=0.5, key='P_mean_bar')
    st.number_input("Frequency [Hz]", 1, 100, step=1, key='f')

    st.subheader("Losses")
    flow_loss    = st.checkbox("Regen flow loss (Ergun)", True)
    regen_imp    = st.checkbox("Regen imperfection",      True)
    mech_loss    = st.checkbox("Mech friction",           True)
    wall_cond    = st.checkbox("Wall conduction",         True)
    leak_loss    = st.checkbox("Seal leakage",            True)
    shuttle_loss = st.checkbox("Shuttle heat loss",       True)
    st.slider("ε_reg",  0.5, 0.99, step=0.01, key='eps_reg')
    st.slider("η_mech", 0.5, 0.99, step=0.01, key='eta_mech')
    st.slider("C_leak", 0.0, 0.20, step=0.01, key='C_leak')

    st.subheader("Shuttle Loss Geometry")
    st.number_input("Displacer radial gap [mm]", 0.10, 0.50, step=0.05, key='gap_mm')
    st.number_input("Effective displacer length [mm]", 50.0, 235.0, step=5.0,
                    key='L_displacer_effective_mm')


# ── Build shared params ───────────────────────────────────────────────────────
params = {k: st.session_state[k] for k in PROTOTYPE if k in st.session_state}
if 'T_h' not in params:
    params['T_h'] = PROTOTYPE['T_h']
for k in ('V_loop_cold', 'V_loop_hot', 'V_cle', 'V_clc',
          'L_displacer', 'P_ref', 'k_metal', 't_wall',
          'D_displacer_effective'):
    params[k] = PROTOTYPE[k]
params['gap']                   = st.session_state.get('gap_mm',
                                    PROTOTYPE['gap'] * 1e3) * 1e-3
params['L_displacer_effective'] = st.session_state.get('L_displacer_effective_mm',
                                    PROTOTYPE['L_displacer_effective'] * 1e3) * 1e-3

losses_flags = dict(flow=flow_loss, regen_imp=regen_imp, mechanical=mech_loss,
                    wall_cond=wall_cond, leakage=leak_loss, shuttle=shuttle_loss)

fixed_Qin_mode = (st.session_state.get('driving_mode', 'Fixed Hot Temperature (T_h)')
                  == "Fixed Heat Input (Q_in)")
Q_in_max = st.session_state.get('Q_in_max', 1000)

# ── Run simulations ───────────────────────────────────────────────────────────
with st.spinner("Computing..."):
    if fixed_Qin_mode:
        sim_s = simulate_fixed_heat(params, Q_in_max, model='schmidt',
                                    losses_flags=losses_flags)
        params_noreg = dict(params); params_noreg['L_r'] = 0.001
        sim_s_noreg  = simulate_fixed_heat(params_noreg, Q_in_max, model='schmidt',
                                           losses_flags=losses_flags)
        sim_a        = simulate_fixed_heat(params, Q_in_max, model='adiabatic',
                                           losses_flags=losses_flags)
    else:
        sim_s = simulate(params, model='schmidt', losses_flags=losses_flags)
        params_noreg = dict(params); params_noreg['L_r'] = 0.001
        sim_s_noreg  = simulate(params_noreg, model='schmidt', losses_flags=losses_flags)
        sim_a        = simulate(params, model='adiabatic', losses_flags=losses_flags)

# Store baseline (Prototype 1) for optimization stage comparisons
if sim_s is not None and 'baseline' not in st.session_state:
    st.session_state['baseline'] = sim_s


# ── Dead-volume helper ────────────────────────────────────────────────────────
def _dead_vol(sim):
    g = sim['geom']
    return (g.get('V_k', 0) + g.get('V_r', 0) + g.get('V_h', 0)) * 1e6


# ── Heat budget helper ────────────────────────────────────────────────────────
def show_heat_budget(losses, Q_in_max, fixed_Qin_mode=False):
    Q_req    = losses['Q_in_W']
    margin   = Q_in_max - Q_req
    feasible = margin >= -0.1

    if fixed_Qin_mode:
        st.success(f"✅ Heat target matched perfectly. Engine consumes {Q_req:.1f} W.")
    else:
        if feasible:
            st.success(f"✅ Heat source is sufficient. Surplus: {margin:.1f} W")
        else:
            st.error(f"❌ Heat source is insufficient. Deficit: {abs(margin):.1f} W")

    st.table({
        "Parameter": [
            "Required heat input (Est. Upper-Bound) [W]",
            "Available / Target heat input [W]",
            "Heat surplus / deficit [W]",
            "Feasible?",
            "Brake power [W]",
        ],
        "Value": [
            f"{Q_req:.1f}",
            f"{Q_in_max:.1f}",
            f"{margin:+.1f}",
            "Yes" if feasible else "No",
            f"{losses['P_brake']:.2f}",
        ]
    })
    st.caption(
        "⚠️ Q_in is an estimated upper bound (Q_miss uses total gas mass). "
        "Actual heater power will be higher."
    )
    if not fixed_Qin_mode:
        st.info(
            "📌 The Heat Budget acts as a constraint check only. "
            "Physically, a large surplus causes T_h to rise — increase the T_h slider to model this."
        )


# ── Tabs ──────────────────────────────────────────────────────────────────────
tab_schmidt, tab_adiabatic, tab_both, tab_optimize = st.tabs([
    "📊 Schmidt (Isothermal)",
    "🔁 Adiabatic (RK45)",
    "⚖️ Both Models",
    "🎯 Optimization",
])


# ════════════════════════════════════════════════════════════════════════════
# TAB 1 — SCHMIDT
# ════════════════════════════════════════════════════════════════════════════
with tab_schmidt:
    st.header("📊 Schmidt (Isothermal) Model")

    use_regen = st.checkbox("With regenerator / heat recovery", value=True,
                            key='schmidt_regen')
    sim = sim_s if use_regen else sim_s_noreg

    if sim is None:
        st.error("❌ Simulation failed — check parameters. "
                 "In Fixed Q_in mode, try a value between 50–800 W.")
    else:
        L = sim['losses']

        # 1. Mode banner
        if fixed_Qin_mode:
            T_h_solved = sim.get('T_h_solved', sim['gas']['T_h'])
            st.success(
                f"✅ **Fixed Heat Input mode** — Q_in target = **{Q_in_max} W**  →  "
                f"Solved T_h = **{T_h_solved:.1f} K** "
                f"(converged in {sim.get('solver_iters', '?')} iterations)"
            )

        # 2. Top Metrics
        c0, c1, c2, c3, c4 = st.columns(5)
        c0.metric("W_cycle [J]",    f"{L['W_cycle']:.4f}")
        c1.metric("W_shaft [J]",    f"{L['W_shaft']:.4f}")
        c2.metric("Brake Power [W]",f"{L['P_brake']:.2f}")
        c3.metric("η_brake [%]",    f"{L['eta_brake']*100:.2f}")
        _fqm = (st.session_state.get('driving_mode') == "Fixed Heat Input (Q_in)")
        if _fqm and 'T_e_max' in L:
            c4.metric("🔥 Calculated T_h [K]", f"{sim['gas']['T_h']:.1f}")

        # 3. Detailed Results Table
        st.subheader("📋 Detailed Results")
        st.table({
            "Parameter": [
                "Mass (gas mass)", "Mean Pressure",
                "W_cycle", "W_shaft", "Brake Power",
                "Q_in [Upper-Bound]", "Q_miss [Upper-Bound]",
                "Q_shuttle", "Efficiency", "% Carnot",
            ],
            "Value": [
                f"{L['M']*1000:.4f} g",
                f"{L['P_mean']/1e5:.3f} bar",
                f"{L['W_cycle']:.4f} J",
                f"{L['W_shaft']:.4f} J",
                f"{L['P_brake']:.2f} W",
                f"{L['Q_in_W']:.2f} W (est. upper-bound)",
                f"{L['Q_miss']*params['f']:.2f} W (est. upper-bound)",
                f"{L.get('Q_shuttle_W', 0):.2f} W",
                f"{L['eta_brake']*100:.3f} %",
                f"{L['frac_carnot']*100:.2f} %",
            ],
        })

        # 4. Diagrams — selected configuration only
        st.subheader("📈 Diagrams")
        if sim is not None:
            R     = sim['result']
            dv    = _dead_vol(sim)
            V_tot = (R['V_e'] + R['V_c']) * 1e6 + dv
            th    = np.rad2deg(R['theta'])

            fig, axes = plt.subplots(1, 2, figsize=(13, 5))
            fig.patch.set_facecolor('white')

            ax = axes[0]
            ax.set_facecolor('white')
            ax.plot(V_tot, R['P'] / 1e5, '#1565C0', lw=2, label='Selected configuration')
            ax.fill(V_tot, R['P'] / 1e5, alpha=0.10, color='#1565C0')
            ax.set(xlabel='V_total [cm³]', ylabel='P [bar]', title='P-V Diagram')
            ax.legend(fontsize=8); ax.grid(alpha=0.3)

            ax = axes[1]
            ax.set_facecolor('white')
            ax.plot(th, R['P'] / 1e5, '#1565C0', lw=2)
            ax.set(xlabel='θ [°]', ylabel='P [bar]', title='Pressure vs θ')
            ax.grid(alpha=0.3)

            plt.tight_layout()
            st.pyplot(fig)
            plt.close(fig)

        # 5. Heat Input Budget
        st.subheader("🔥 Heat Input Budget — Estimated Upper-Bound")
        show_heat_budget(L, Q_in_max, fixed_Qin_mode=_fqm)

        # 6. Nonlinear Q_in Sweep
        st.subheader("🔥 Q_in Sweep — With vs Without Regenerator")
        st.caption(
            "For each Q_in value, the model solves for T_h from the energy balance. "
            "This is a nonlinear result — not a linear scaling. "
            "Q_in is net heat into the gas. Actual heater power is higher."
        )
        with st.spinner("Running Q_in sweep..."):
            Q_in_range = np.linspace(50, Q_in_max, 40)
            lf_sweep = dict(flow=True, regen_imp=True, mechanical=True,
                            wall_cond=True, leakage=True, shuttle=False)
            sw_with  = sweep_qin(params, Q_in_range, 'schmidt', lf_sweep, no_regen=False)
            sw_noreg = sweep_qin(params, Q_in_range, 'schmidt', lf_sweep, no_regen=True)

        if len(sw_with.get('Q_in', [])) > 0 and len(sw_noreg.get('Q_in', [])) > 0:
            fig_sw, axes_sw = plt.subplots(1, 2, figsize=(13, 5))
            fig_sw.patch.set_facecolor('white')

            ax = axes_sw[0]
            ax.set_facecolor('white')
            ax.plot(sw_with['Q_in'],  sw_with['P_brake'],  '#1565C0', lw=2,
                    label='With regenerator')
            ax.plot(sw_noreg['Q_in'], sw_noreg['P_brake'], '#C62828', lw=2,
                    ls='--', label='Without regenerator')
            ax.axvline(Q_in_max, color='gray', ls=':', lw=1.2,
                       label=f'Q_in_max = {Q_in_max} W')
            ax.set(xlabel='Q_in [W]', ylabel='P_brake [W]',
                   title='Brake Power vs Heat Input')
            ax.legend(); ax.grid(alpha=0.3)

            ax = axes_sw[1]
            ax.set_facecolor('white')
            ax.plot(sw_with['Q_in'],  sw_with['eta_net']  * 100, '#1565C0', lw=2,
                    label='With regenerator')
            ax.plot(sw_noreg['Q_in'], sw_noreg['eta_net'] * 100, '#C62828', lw=2,
                    ls='--', label='Without regenerator')
            ax.axvline(Q_in_max, color='gray', ls=':', lw=1.2)
            ax.set(xlabel='Q_in [W]', ylabel='η_net [%]',
                   title='Net Efficiency vs Heat Input')
            ax.legend(); ax.grid(alpha=0.3)

            plt.tight_layout()
            st.pyplot(fig_sw)
            plt.close(fig_sw)
        else:
            st.warning("Q_in sweep produced no results — check parameters.")

        # 7. Engine Animation
        st.subheader("🎬 Engine Animation")
        if st.button("Generate Animation", key='anim_s'):
            with st.spinner("Rendering..."):
                try:
                    geom_frozen   = _freeze_dict(build_geometry(to_si(params)))
                    params_frozen = _freeze_dict(params)
                    gif_b64 = build_engine_animation(geom_frozen, params_frozen)
                    st.markdown(
                        f'<img src="data:image/gif;base64,{gif_b64}" '
                        f'style="width:100%;max-width:780px;border-radius:8px;" />',
                        unsafe_allow_html=True,
                    )
                except Exception as e:
                    st.warning(f"Animation error: {e}")

        # 8. Top-3 Geometric Improvements
        st.subheader("🔧 Top-3 Geometric Improvements (Global Sweep)")
        st.caption("Full-range sweep per geometric parameter — operating conditions fixed.")
        with st.spinner("Running global sensitivity analysis..."):
            sens = geometry_sensitivity(params, losses_flags)
        if sens:
            for rank, (key, name, units, base_val, best_val, delta_W, pct) in \
                    enumerate(sens[:3], 1):
                direction = "↑ Increase" if best_val > base_val else "↓ Decrease"
                arrow = "🟢" if delta_W > 0 else "🔴"
                with st.expander(
                    f"#{rank}  {name}  —  {arrow} {delta_W:+.2f} W  ({pct:+.1f}%)",
                    expanded=(rank == 1)
                ):
                    st.markdown(
                        f"**{direction} {name}** from **{base_val:.2f} {units}** "
                        f"to **{best_val:.2f} {units}**\n\n"
                        f"Estimated brake power change: **{delta_W:+.2f} W** ({pct:+.1f}%)\n\n"
                        f"*Full-range sweep (Schmidt model, all other params fixed).*"
                    )
        else:
            st.info("Could not compute sensitivity — check parameters.")

        # 9. Power vs Mean Pressure sweep
        st.subheader("📉 Power vs Mean Pressure (All Gases)")
        st.caption("Quick parametric sweep — all other parameters held constant.")
        with st.spinner("Running pressure sweep..."):
            lf_ps = dict(flow=flow_loss, regen_imp=regen_imp, mechanical=mech_loss,
                         wall_cond=wall_cond, leakage=leak_loss, shuttle=shuttle_loss)
            P_range = np.linspace(0.5, min(30.0, params['P_mean_bar'] * 4), 20)
            gas_colors = {'Air': '#1565C0', 'Helium': '#C62828', 'Hydrogen': '#2E7D32'}

            fig_ps, ax_ps = plt.subplots(figsize=(9, 5))
            fig_ps.patch.set_facecolor('white'); ax_ps.set_facecolor('white')
            for gas_name, gcolor in gas_colors.items():
                pw_vals = []
                for pm in P_range:
                    p_sw = dict(params); p_sw['gas'] = gas_name
                    p_sw['P_mean_bar'] = float(pm)
                    s_sw = simulate(p_sw, model='schmidt', losses_flags=lf_ps)
                    pw_vals.append(s_sw['losses']['P_brake'] if s_sw else float('nan'))
                ax_ps.plot(P_range, pw_vals, color=gcolor, lw=2, label=gas_name)
            ax_ps.axvline(params['P_mean_bar'], color='#888', ls='--', lw=1.2,
                          label=f"Current = {params['P_mean_bar']:.1f} bar")
            ax_ps.set(xlabel='Mean Pressure [bar]', ylabel='Brake Power [W]',
                      title='Brake Power vs Mean Pressure — Schmidt')
            ax_ps.legend(fontsize=9); ax_ps.grid(alpha=0.3)
            plt.tight_layout()
            st.pyplot(fig_ps); plt.close(fig_ps)

        # 10. Export
        if st.button("📥 Export to Excel (Schmidt)"):
            try:
                excel_bytes = export_excel_v10("Schmidt", sim, Q_in_max)
                fname = f"stirling_v10_schmidt_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                st.download_button("⬇️ Download Excel (Schmidt)", excel_bytes, fname,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            except Exception as e:
                st.error(f"Export error: {e}")


# ════════════════════════════════════════════════════════════════════════════
# TAB 2 — ADIABATIC
# ════════════════════════════════════════════════════════════════════════════
with tab_adiabatic:
    st.header("🔁 Adiabatic (RK45) Model")

    if sim_a is None:
        st.error("❌ Adiabatic simulation failed — check parameters.")
    else:
        La = sim_a['losses']
        Ls = sim_s['losses'] if sim_s else None
        _fqm_a = (st.session_state.get('driving_mode') == "Fixed Heat Input (Q_in)")

        # 1. Mode banner
        if _fqm_a:
            T_h_solved_a = sim_a.get('T_h_solved', sim_a['gas']['T_h'])
            st.success(
                f"✅ **Fixed Heat Input mode** — Q_in target = **{Q_in_max} W**  →  "
                f"Solved T_h = **{T_h_solved_a:.1f} K** "
                f"(converged in {sim_a.get('solver_iters', '?')} iterations)"
            )

        # 2. Top Metrics
        c0, c1, c2, c3, c4 = st.columns(5)
        c0.metric("W_cycle [J]",    f"{La['W_cycle']:.4f}")
        c1.metric("W_shaft [J]",    f"{La['W_shaft']:.4f}")
        c2.metric("Brake Power [W]",f"{La['P_brake']:.2f}")
        c3.metric("η_brake [%]",    f"{La['eta_brake']*100:.2f}")
        if _fqm_a and 'T_e_max' in La:
            c4.metric("🔥 Calculated T_h [K]", f"{sim_a['gas']['T_h']:.1f}")

        # 3. Detailed Results Table
        st.subheader("📋 Detailed Results")
        st.table({
            "Parameter": [
                "Mass (gas mass)", "Mean Pressure",
                "W_cycle", "W_shaft", "Brake Power",
                "Q_in [Upper-Bound]", "Q_miss [Upper-Bound]",
                "Q_shuttle", "Efficiency", "% Carnot",
            ],
            "Value": [
                f"{La['M']*1000:.4f} g",
                f"{La['P_mean']/1e5:.3f} bar",
                f"{La['W_cycle']:.4f} J",
                f"{La['W_shaft']:.4f} J",
                f"{La['P_brake']:.2f} W",
                f"{La['Q_in_W']:.2f} W (est. upper-bound)",
                f"{La['Q_miss']*params['f']:.2f} W (est. upper-bound)",
                f"{La.get('Q_shuttle_W', 0):.2f} W",
                f"{La['eta_brake']*100:.3f} %",
                f"{La['frac_carnot']*100:.2f} %",
            ],
        })

        if Ls:
            t_ratio  = params['T_h'] / params['T_k']
            base_msg = (
                f"⚖️ M = {Ls['M']*1000:.4f} g (Schmidt). "
                f"Adiabatic P_mean = {La['P_mean']/1e5:.3f} bar (output). "
                f"T_h/T_k = {t_ratio:.2f}."
            )
            if La['W_cycle'] > Ls['W_cycle']:
                st.warning(base_msg + f" ⚠️ W_adiabatic > W_schmidt — validate results.")
            else:
                st.info(base_msg)

        # 4. Diagrams
        st.subheader("📈 Diagrams")
        if sim_a is not None:
            Ra    = sim_a['result']
            dv_a  = _dead_vol(sim_a)
            Va    = (Ra['V_e'] + Ra['V_c']) * 1e6 + dv_a
            th_a  = np.rad2deg(Ra['theta'])

            fig, axes = plt.subplots(1, 3, figsize=(16, 5))
            fig.patch.set_facecolor('white')

            axes[0].set_facecolor('white')
            axes[0].plot(Va, Ra['P'] / 1e5, '#C62828', lw=2)
            axes[0].fill(Va, Ra['P'] / 1e5, alpha=0.10, color='#C62828')
            axes[0].set(xlabel='V_total [cm³]', ylabel='P [bar]', title='P-V Diagram')
            axes[0].grid(alpha=0.3)

            axes[1].set_facecolor('white')
            axes[1].plot(th_a, Ra['P'] / 1e5, '#C62828', lw=2)
            axes[1].set(xlabel='θ [°]', ylabel='P [bar]', title='Pressure vs θ')
            axes[1].grid(alpha=0.3)

            axes[2].set_facecolor('white')
            axes[2].axhline(params['T_h'], color='#C62828', ls='--', lw=1.5,
                            label=f'T_h = {params["T_h"]} K')
            axes[2].axhline(params['T_k'], color='#1565C0', ls='--', lw=1.5,
                            label=f'T_k = {params["T_k"]} K')
            axes[2].plot(th_a, Ra['T_e'], '#EF5350', lw=2, label='T_e')
            axes[2].plot(th_a, Ra['T_c'], '#42A5F5', lw=2, label='T_c')
            axes[2].set(xlabel='θ [°]', ylabel='T [K]', title='Gas Temperatures')
            axes[2].legend(fontsize=8); axes[2].grid(alpha=0.3)

            plt.tight_layout()
            st.pyplot(fig)
            plt.close(fig)

        # 5. Heat Input Budget
        st.subheader("🔥 Heat Input Budget — Estimated Upper-Bound (Adiabatic)")
        show_heat_budget(La, Q_in_max, fixed_Qin_mode=_fqm_a)

        # 6. Validation checks
        with st.expander("🔍 Validation checks", expanded=False):
            delta, ok = validate_mass_conservation(sim_a['result'], sim_a['geom'], sim_a['gas'])
            (st.success if ok else st.error)(
                f"{'✅' if ok else '❌'} Mass conservation: {delta:.3f} %"
                + (" < 2 % ✓" if ok else " > 2 % limit"))
            err, ok = validate_first_law(La)
            (st.success if ok else st.error)(
                f"{'✅' if ok else '❌'} First-law error: {err:.6f} %")
            eta_b, eta_c, ok = validate_carnot(La)
            (st.success if ok else st.error)(
                f"{'✅' if ok else '❌'} η_brake = {eta_b*100:.3f}%  "
                f"{'≤' if ok else '>'} η_Carnot = {eta_c*100:.2f}%  "
                f"({La['frac_carnot']*100:.1f}% of Carnot)")
            W1, W2, ratio_p, ok_p = validate_pressure_scaling(params, losses_flags)
            if W1:
                (st.success if ok_p else st.error)(
                    f"{'✅' if ok_p else '❌'} Pressure scaling ratio = {ratio_p:.5f} (expected 2.000)")

        # 7. Export
        if st.button("📥 Export to Excel (Adiabatic)"):
            try:
                excel_bytes = export_excel_v10("Adiabatic", sim_a, Q_in_max)
                fname = f"stirling_v10_adiabatic_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                st.download_button("⬇️ Download Excel (Adiabatic)", excel_bytes, fname,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            except Exception as e:
                st.error(f"Export error: {e}")


# ════════════════════════════════════════════════════════════════════════════
# TAB 3 — BOTH MODELS
# ════════════════════════════════════════════════════════════════════════════
with tab_both:
    st.header("⚖️ Both Models — Side-by-Side Comparison")

    use_regen_both = st.checkbox("Schmidt: with regenerator", value=True,
                                 key='both_regen')
    sim_s_display = sim_s if use_regen_both else sim_s_noreg

    if sim_s_display and sim_a:
        Ls = sim_s_display['losses']
        La = sim_a['losses']
        _fqm_b = (st.session_state.get('driving_mode') == "Fixed Heat Input (Q_in)")

        if _fqm_b:
            Th_s = sim_s_display.get('T_h_solved', sim_s_display['gas']['T_h'])
            Th_a = sim_a.get('T_h_solved', sim_a['gas']['T_h'])
            st.info(
                f"🔁 **Fixed Heat Input mode** — Q_in target = **{Q_in_max} W**\n\n"
                f"Schmidt solved T_h = **{Th_s:.1f} K** | "
                f"Adiabatic solved T_h = **{Th_a:.1f} K**"
            )

        st.subheader("Comparison Table")
        st.table({
            "Metric":     ["M [g]","W_cycle [J]","W_shaft [J]","P_brake [W]",
                           "η_brake [%]","P_mean_out [bar]"],
            "Schmidt":    [f"{Ls['M']*1000:.4f}", f"{Ls['W_cycle']:.4f}",
                           f"{Ls['W_shaft']:.4f}", f"{Ls['P_brake']:.2f}",
                           f"{Ls['eta_brake']*100:.2f}", f"{Ls['P_mean']/1e5:.3f}"],
            "Adiabatic":  [f"{La['M']*1000:.4f}", f"{La['W_cycle']:.4f}",
                           f"{La['W_shaft']:.4f}", f"{La['P_brake']:.2f}",
                           f"{La['eta_brake']*100:.2f}", f"{La['P_mean']/1e5:.3f}"],
        })

        st.subheader("📈 Overlaid Comparison Plots")
        Rs   = sim_s_display['result']
        Ra   = sim_a['result']
        V_s  = (Rs['V_e'] + Rs['V_c']) * 1e6 + _dead_vol(sim_s_display)
        V_a  = (Ra['V_e'] + Ra['V_c']) * 1e6 + _dead_vol(sim_a)
        th_a = np.rad2deg(Ra['theta'])

        fig_b, axes_b = plt.subplots(1, 2, figsize=(14, 5))
        fig_b.patch.set_facecolor('white')

        ax = axes_b[0]
        ax.set_facecolor('white')
        ax.plot(V_s, Rs['P'] / 1e5, '#1565C0', lw=2, label='Selected configuration')
        ax.fill(V_s, Rs['P'] / 1e5, alpha=0.08, color='#1565C0')
        ax.plot(V_a, Ra['P'] / 1e5, '#C62828', lw=2, ls='--', label='Adiabatic')
        ax.fill(V_a, Ra['P'] / 1e5, alpha=0.08, color='#C62828')
        ax.set(xlabel='V_total [cm³]', ylabel='P [bar]', title='P-V Diagram — Schmidt vs Adiabatic')
        ax.legend(fontsize=9); ax.grid(alpha=0.3)

        ax = axes_b[1]
        ax.set_facecolor('white')
        ax.plot(th_a, Ra['T_e'], '#D32F2F', lw=2, label='T_e (Adiabatic)')
        ax.plot(th_a, Ra['T_c'], '#1976D2', lw=2, label='T_c (Adiabatic)')
        ax.axhline(params['T_h'], color='#D32F2F', ls='--', lw=1.5, alpha=0.55,
                   label=f"T_h = {params['T_h']} K (wall)")
        ax.axhline(params['T_k'], color='#1976D2', ls='--', lw=1.5, alpha=0.55,
                   label=f"T_k = {params['T_k']} K (wall)")
        ax.set(xlabel='θ [°]', ylabel='T [K]', title='Gas Temperatures — Adiabatic vs Walls')
        ax.legend(fontsize=8); ax.grid(alpha=0.3)

        plt.tight_layout()
        st.pyplot(fig_b); plt.close(fig_b)

        st.subheader("🔥 Heat Input Budget")
        c1, c2 = st.columns(2)
        with c1:
            st.markdown("**Schmidt**")
            show_heat_budget(Ls, Q_in_max, fixed_Qin_mode=_fqm_b)
        with c2:
            st.markdown("**Adiabatic**")
            show_heat_budget(La, Q_in_max, fixed_Qin_mode=_fqm_b)

        if st.button("📥 Export to Excel (Both Models)"):
            try:
                excel_bytes = export_excel_v10("Schmidt", sim_s_display, Q_in_max)
                fname = f"stirling_v10_both_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                st.download_button("⬇️ Download Excel (Both)", excel_bytes, fname,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            except Exception as e:
                st.error(f"Export error: {e}")
    else:
        st.error("One or both simulations failed — check parameters.")


# ════════════════════════════════════════════════════════════════════════════
# TAB 4 — OPTIMIZATION
# ════════════════════════════════════════════════════════════════════════════
with tab_optimize:
    st.header("🎯 Optimization")

    obj_map = {"Max Power": "power", "Max Efficiency": "efficiency",
               "Balanced": "balanced"}

    stage_choice = st.radio(
        "Select optimization stage:",
        ["Stage 2 — Prototype 2 (locked displacer)",
         "Stage 3 — Full Geometry",
         "Stage 4 — Operating Conditions (requires Stage 2 first)"],
        horizontal=False
    )

    # ── Stage 2 ───────────────────────────────────────────────────────────────
    if stage_choice.startswith("Stage 2"):
        st.subheader("Stage 2 — Prototype 2")
        st.info(
            "🔒 **Locked:** D_displacer, S_displacer, L_displacer "
            "(existing physical displacer). "
            "**Open:** everything else listed below."
        )

        stage2_open_defs = [
            ('D_power',   'Power piston diameter', 30,  120, 'mm'),
            ('S_power',   'Power piston stroke',   20,  120, 'mm'),
            ('phi_deg',   'Phase angle',           60,  120, '°'),
            ('D_r',       'Regen diameter',        20,   80, 'mm'),
            ('L_r',       'Regen length',          50,  400, 'mm'),
            ('d_wire',    'Wire diameter',         0.5,  3.0, 'mm'),
            ('porosity',  'Porosity',              0.5, 0.99, ''),
        ]

        open_specs_s2 = []
        cols_s2 = st.columns(2)
        for idx, (key, name, lo, hi, unit) in enumerate(stage2_open_defs):
            with cols_s2[idx % 2]:
                if st.checkbox(f"Open: {name} ({lo}–{hi} {unit})",
                               value=True, key=f"s2_{key}"):
                    open_specs_s2.append((key, lo, hi))

        obj_s2  = st.radio("Objective:", ["Max Power", "Max Efficiency", "Balanced"],
                           horizontal=True, key="obj_s2")
        method_s2 = st.radio("Search method:",
                             ["Coarse-Fine (fast)", "LHS (better coverage)"],
                             horizontal=True, key="method_s2")

        if st.button("🚀 Run Prototype 2 Optimization",
                     disabled=len(open_specs_s2) == 0, key="run_s2"):
            lf_s2 = dict(flow=True, regen_imp=True, mechanical=True,
                         wall_cond=True, leakage=True, shuttle=False)
            p_s2  = dict(params); p_s2['Q_in_max'] = Q_in_max

            prog2 = st.progress(0, "Running Prototype 2 search...")
            def cb_s2(f, msg): prog2.progress(min(f, 1.0), msg)

            method_key = 'lhs' if 'LHS' in method_s2 else 'coarse_fine'
            try:
                bp2, bl2, all2 = prototype2_search(
                    p_s2, open_specs_s2, obj_map[obj_s2],
                    'schmidt', lf_s2, method=method_key,
                    n_samples=300, progress_cb=cb_s2)
                prog2.empty()

                if bl2 is None:
                    st.error("Optimization failed — no feasible solution found.")
                else:
                    st.session_state['proto2_result'] = {
                        'params': bp2, 'losses': bl2, 'all_results': all2,
                        'geom': build_geometry(to_si(bp2))}

                    q2 = bl2['Q_in_W']
                    if q2 <= Q_in_max:
                        st.success(f"✅ Feasible — Q_in = {q2:.1f} W ≤ {Q_in_max} W")
                    else:
                        st.error(f"❌ Infeasible — Q_in = {q2:.1f} W > {Q_in_max} W")

                    ca, cb_c, cc, cd = st.columns(4)
                    ca.metric("P_brake [W]", f"{bl2['P_brake']:.2f}")
                    cb_c.metric("η_net [%]",  f"{bl2['eta_brake']*100:.2f}")
                    cc.metric("Q_in [W]",    f"{q2:.1f}")
                    cd.metric("M [g]",       f"{bl2['M']*1000:.3f}")

                    # V_swc scatter
                    V_swc_all, P_all2 = [], []
                    for _, p_i, l_i in all2:
                        D = p_i.get('D_power', 65.6) * 1e-3
                        S = p_i.get('S_power',  61.6) * 1e-3
                        V_swc_all.append(math.pi * (D/2)**2 * S * 1e6)
                        P_all2.append(l_i['P_brake'])

                    if V_swc_all:
                        fig_v, ax_v = plt.subplots(figsize=(8, 4))
                        fig_v.patch.set_facecolor('white')
                        ax_v.scatter(V_swc_all, P_all2, alpha=0.4, s=18, color='#888')
                        D_best = bp2.get('D_power', 65.6) * 1e-3
                        S_best = bp2.get('S_power',  61.6) * 1e-3
                        V_best = math.pi * (D_best/2)**2 * S_best * 1e6
                        ax_v.scatter([V_best], [bl2['P_brake']], s=120, color='#C62828',
                                     zorder=5, label=f'Best: {V_best:.1f} cm³')
                        ax_v.set(xlabel='V_swc [cm³]', ylabel='P_brake [W]',
                                 title='P_brake vs Power Piston Swept Volume')
                        ax_v.legend(); ax_v.grid(alpha=0.3)
                        st.pyplot(fig_v); plt.close(fig_v)
            except Exception as e:
                prog2.empty()
                st.error(f"Stage 2 error: {e}")

        # Show saved Stage 2 result
        if 'proto2_result' in st.session_state:
            r2 = st.session_state['proto2_result']
            bp2 = r2['params']; bl2 = r2['losses']
            st.info(
                f"💾 Saved Prototype 2: D_power={bp2.get('D_power',65.6):.1f} mm, "
                f"S_power={bp2.get('S_power',61.6):.1f} mm, "
                f"φ={bp2.get('phi_deg',90):.0f}°, "
                f"P_brake={bl2['P_brake']:.2f} W"
            )

    # ── Stage 3 ───────────────────────────────────────────────────────────────
    elif stage_choice.startswith("Stage 3"):
        st.subheader("Stage 3 — Full Geometry Optimization")
        st.info("Gas = Air, P = 1 bar, f = 10 Hz. Q_in ≤ 1500 W. All geometry open.")

        stage3_open_defs = [
            ('D_displacer', 'Displacer diameter', 40,  120, 'mm'),
            ('S_displacer', 'Displacer stroke',   50,  180, 'mm'),
            ('D_power',     'Power piston diam',  30,  120, 'mm'),
            ('S_power',     'Power piston stroke', 20, 120, 'mm'),
            ('phi_deg',     'Phase angle',         60, 120, '°'),
            ('D_r',         'Regen diameter',      20,  80, 'mm'),
            ('L_r',         'Regen length',        50, 400, 'mm'),
            ('d_wire',      'Wire diameter',       0.5, 3.0, 'mm'),
        ]

        open_specs_s3 = []
        cols_s3 = st.columns(2)
        for idx, (key, name, lo, hi, unit) in enumerate(stage3_open_defs):
            with cols_s3[idx % 2]:
                if st.checkbox(f"Open: {name} ({lo}–{hi} {unit})",
                               value=True, key=f"s3_{key}"):
                    open_specs_s3.append((key, lo, hi))

        st.markdown("### Balanced score")
        st.caption(
            "Stage 3 always returns three named designs: "
            "Max Power, Max Efficiency, and Balanced. "
            "This slider affects only the Balanced point."
        )
        w_P_s3 = st.slider(
            "Balanced weight: Power ↔ Efficiency",
            0.0, 1.0, 0.5, 0.05,
            key="wP_s3"
        )
        st.caption(
            f"Balanced Score = {w_P_s3:.2f} × (P/P_ref) + "
            f"{1-w_P_s3:.2f} × (η/η_ref)"
        )

        method_s3 = st.radio("Search method:",
                             ["LHS (recommended)", "Coarse-Fine", "Bayesian"],
                             horizontal=True, key="method_s3")
        n_s3 = (st.slider("LHS samples", 100, 1000, 400, 50, key="n_s3")
                if "LHS" in method_s3 else 400)

        Q_in_max_s3 = st.slider("Q_in ceiling [W]", 500, 3000, 1500, 100,
                                 key="qmax_s3")

        if st.button("🚀 Run Stage 3 Optimization",
                     disabled=len(open_specs_s3) == 0, key="run_s3"):
            lf_s3 = dict(flow=True, regen_imp=True, mechanical=True,
                         wall_cond=True, leakage=True, shuttle=False)
            p_s3  = dict(params)
            p_s3.update({'gas': 'Air', 'P_mean_bar': 1.0,
                         'f': 10.0, 'Q_in_max': Q_in_max_s3})

            prog3 = st.progress(0, "Running Stage 3...")
            def cb_s3(f, msg): prog3.progress(min(f, 1.0), msg)

            method_map = {"LHS (recommended)": "lhs",
                          "Coarse-Fine": "coarse_fine",
                          "Bayesian": "bayesian"}
            try:
                # V10: Stage 3 returns three named feasible designs:
                # Max Power, Max Efficiency, Balanced.
                base_ref = st.session_state.get('baseline', {}).get('losses')
                s3_named = stage3_search_named(
                    p_s3, open_specs_s3, Q_in_max_s3,
                    model_key='schmidt',
                    losses_flags=lf_s3,
                    method=method_map[method_s3],
                    n_samples=n_s3,
                    progress_cb=cb_s3,
                    w_P=w_P_s3,
                    ref_losses=base_ref
                )
                prog3.empty()

                if not s3_named or s3_named.get('balanced') is None:
                    st.error("Stage 3 found no feasible solution.")
                else:
                    st.session_state['stage3_results'] = {
                        'named': s3_named,
                        'best_params': s3_named['balanced']['params'],
                        'best_losses': s3_named['balanced']['losses'],
                        'feasible': s3_named.get('feasible_results', []),
                        'raw': s3_named.get('raw_results', []),
                        'objective': 'Named designs',
                        'Q_in_max': Q_in_max_s3
                    }

                    st.success("✅ Stage 3 complete — Max Power, Max Efficiency, and Balanced designs selected.")

                    # Show the three named designs
                    rows3 = []
                    for label, key in [
                        ("Max Power", "max_power"),
                        ("Max Efficiency", "max_efficiency"),
                        ("Balanced", "balanced"),
                    ]:
                        d = s3_named.get(key)
                        if not d:
                            continue
                        L = d['losses']
                        q = L['Q_in_W']
                        rows3.append({
                            "Design": label,
                            "P_brake [W]": f"{L['P_brake']:.2f}",
                            "η_net [%]": f"{L['eta_brake']*100:.2f}",
                            "Q_in [W]": f"{q:.1f}",
                            "M [g]": f"{L['M']*1000:.3f}",
                            "Feasible": "🟢 Yes" if q <= Q_in_max_s3 else "🔴 No",
                        })
                    if rows3:
                        st.table(rows3)
            except Exception as e:
                prog3.empty()
                st.error(f"Stage 3 error: {e}")

        # Display saved Stage 3 results
        if 'stage3_results' in st.session_state:
            s3r = st.session_state['stage3_results']
            raw3  = s3r.get('raw', [])
            bp3   = s3r.get('best_params')
            bl3   = s3r.get('best_losses')
            Q3max = s3r.get('Q_in_max', 1500)

            # Power–Efficiency map
            st.subheader("📊 Power–Efficiency Map")
            named = {
                'proto1': st.session_state['baseline']['losses']
                          if 'baseline' in st.session_state else None,
                'proto2': st.session_state.get('proto2_result', {}).get('losses'),
                'max_power': bl3,
            }
            fig_map = plot_power_efficiency_map(raw3, named,
                                                title="Stage 3 — Power–Efficiency Map")
            st.pyplot(fig_map); plt.close(fig_map)
            st.caption("Gray dots = all evaluated candidates. "
                       "Named points = selected designs. "
                       "Red border = infeasible (Q_in > budget).")

            # Volume plots
            st.subheader("📐 Volume Analysis")
            if raw3:
                V_total_all, V_swc_all, alpha_all, P_all3 = [], [], [], []
                for _, p_i, l_i in raw3:
                    D_d = p_i.get('D_displacer', 75)   * 1e-3
                    S_d = p_i.get('S_displacer', 101.5) * 1e-3
                    D_p = p_i.get('D_power', 65.6)     * 1e-3
                    S_p = p_i.get('S_power',  61.6)    * 1e-3
                    V_swe = math.pi*(D_d/2)**2*S_d * 1e6
                    V_swc = math.pi*(D_p/2)**2*S_p * 1e6
                    V_total_all.append(V_swe + V_swc)
                    V_swc_all.append(V_swc)
                    alpha_all.append(V_swc / V_swe if V_swe > 0 else 0)
                    P_all3.append(l_i['P_brake'])

                fig_vol, axes_vol = plt.subplots(1, 2, figsize=(13, 5))
                fig_vol.patch.set_facecolor('white')
                axes_vol[0].scatter(V_total_all, P_all3, s=12, alpha=0.4, color='#888')
                axes_vol[0].set(xlabel='V_total_swept [cm³]', ylabel='P_brake [W]',
                                title='P_brake vs Total Swept Volume')
                axes_vol[0].grid(alpha=0.3)
                axes_vol[1].scatter(alpha_all, P_all3, s=12, alpha=0.4, color='#888')
                axes_vol[1].set(xlabel='α = V_swc / V_swe', ylabel='P_brake [W]',
                                title='P_brake vs Volume Ratio α')
                axes_vol[1].grid(alpha=0.3)
                plt.tight_layout()
                st.pyplot(fig_vol); plt.close(fig_vol)
                st.caption("α = V_swc / V_swe. Higher α = larger power piston relative to displacer.")

            # Frequency & Torque Sweep
            st.subheader("⚡ Frequency & Torque Sweep")
            st.caption(
                "Performance sweep at f = 5–25 Hz on selected designs. "
                "Torque = P_brake / (2π·f)."
            )
            f_vals_sweep = np.linspace(5, 25, 20)
            lf_ft = dict(flow=True, regen_imp=True, mechanical=True,
                         wall_cond=True, leakage=True, shuttle=False)
            designs_for_sweep = {}
            if 'baseline' in st.session_state:
                designs_for_sweep['Prototype 1'] = (
                    st.session_state['baseline']['params'], '#1565C0')
            if 'proto2_result' in st.session_state:
                designs_for_sweep['Prototype 2'] = (
                    st.session_state['proto2_result']['params'], '#2E7D32')
            if bp3 is not None:
                designs_for_sweep['Max Power (S3)'] = (bp3, '#C62828')

            if designs_for_sweep:
                fig_ft, axes_ft = plt.subplots(1, 2, figsize=(13, 5))
                fig_ft.patch.set_facecolor('white')
                for label, (p_des, color) in designs_for_sweep.items():
                    ft = frequency_torque_sweep(p_des, f_vals_sweep, 'schmidt', lf_ft)
                    if len(ft.get('f', [])) == 0:
                        continue
                    axes_ft[0].plot(ft['f'], ft['P_brake'], color=color, lw=2, label=label)
                    axes_ft[1].plot(ft['f'], ft['Torque'],  color=color, lw=2, label=label)
                axes_ft[0].set(xlabel='f [Hz]', ylabel='P_brake [W]',
                               title='Brake Power vs Frequency')
                axes_ft[0].legend(); axes_ft[0].grid(alpha=0.3)
                axes_ft[1].set(xlabel='f [Hz]', ylabel='Torque [N·m]',
                               title='Torque vs Frequency')
                axes_ft[1].legend(); axes_ft[1].grid(alpha=0.3)
                plt.tight_layout()
                st.pyplot(fig_ft); plt.close(fig_ft)

            # Geometry Summary Table
            st.subheader("📋 Geometry Summary Table")
            designs_table = {}
            if 'baseline' in st.session_state:
                designs_table['proto1'] = st.session_state['baseline']
            if 'proto2_result' in st.session_state:
                designs_table['proto2'] = st.session_state['proto2_result']
            if bp3 is not None and bl3 is not None:
                designs_table['max_power'] = {
                    'params': bp3, 'losses': bl3,
                    'geom': build_geometry(to_si(bp3))}
            show_geometry_summary_table(designs_table)

    # ── Stage 4 ───────────────────────────────────────────────────────────────
    elif stage_choice.startswith("Stage 4"):
        st.subheader("Stage 4 — Operating Conditions on Prototype 2")

        if 'proto2_result' not in st.session_state:
            st.warning(
                "⚠️ Run Stage 2 first to get the Prototype 2 geometry. "
                "Stage 4 optimizes operating conditions on that design."
            )
            st.stop()

        proto2_params = st.session_state['proto2_result']['params']
        bp2_disp      = st.session_state['proto2_result']
        bl2_disp      = bp2_disp['losses']
        st.info(
            f"🔒 Geometry fixed as Prototype 2: "
            f"D_power={proto2_params.get('D_power',65.6):.1f} mm, "
            f"S_power={proto2_params.get('S_power',61.6):.1f} mm, "
            f"D_r={proto2_params.get('D_r',40):.1f} mm, "
            f"φ={proto2_params.get('phi_deg',90):.0f}°"
        )
        st.warning(
            "⚠️ Hydrogen: high theoretical performance but requires strict "
            "safety, sealing, leakage control, and material compatibility review."
        )

        P_max_s4 = st.slider("Max pressure [bar]", 1.0, 10.0, 10.0, 0.5, key="Pmax_s4")
        f_max_s4 = st.slider("Max frequency [Hz]", 5.0, 25.0, 20.0, 1.0, key="fmax_s4")
        obj_s4   = st.radio("Objective:", ["Max Power", "Max Efficiency", "Balanced"],
                            horizontal=True, key="obj_s4")
        n_s4     = st.slider("LHS samples per gas", 50, 400, 200, 50, key="n_s4")

        if st.button("🚀 Run Stage 4 (Air + Helium + Hydrogen)", key="run_s4"):
            lf_s4 = dict(flow=True, regen_imp=True, mechanical=True,
                         wall_cond=True, leakage=True, shuttle=False)
            prog4 = st.progress(0, "Starting Stage 4...")
            def cb_s4(f, msg): prog4.progress(min(f, 1.0), msg)

            try:
                # V10: Stage 4 runs on Prototype 2 geometry and returns
                # Max Power / Max Efficiency / Balanced for each gas.
                base_ref = st.session_state.get('proto2_result', {}).get('losses')
                s4_results = stage4_search_named(
                    proto2_params,
                    Q_in_max=Q_in_max, P_max=P_max_s4, f_max=f_max_s4,
                    model_key='schmidt', losses_flags=lf_s4,
                    n_samples=n_s4, progress_cb=cb_s4,
                    w_P=0.5,
                    ref_losses=base_ref
                )
                prog4.empty()
                st.session_state['stage4_results'] = s4_results
            except Exception as e:
                prog4.empty()
                st.error(f"Stage 4 error: {e}")

        if 'stage4_results' in st.session_state:
            s4 = st.session_state['stage4_results']

            # Gas Comparison Table — V10 named designs
            st.subheader("Gas Comparison Table")
            rows_s4 = []
            for gas_name in ['Air', 'Helium', 'Hydrogen']:
                r = s4.get(gas_name, {})
                for obj_label, key in [
                    ("Max Power", "max_power"),
                    ("Max Efficiency", "max_efficiency"),
                    ("Balanced", "balanced"),
                ]:
                    d = r.get(key)
                    if not d:
                        continue
                    bl_g = d.get('losses')
                    bp_g = d.get('params')
                    if bl_g and bp_g:
                        q_g = bl_g['Q_in_W']
                        rows_s4.append({
                            'Gas':          gas_name,
                            'Objective':    obj_label,
                            'P_mean [bar]': f"{bp_g.get('P_mean_bar', 1):.1f}",
                            'T_h [K]':      f"{bp_g.get('T_h', 873):.0f}",
                            'f [Hz]':       f"{bp_g.get('f', 10):.1f}",
                            'P_brake [W]':  f"{bl_g['P_brake']:.2f}",
                            'η_net [%]':    f"{bl_g['eta_brake']*100:.2f}",
                            'Q_in [W]':     f"{q_g:.1f}",
                            'Feasible':     '🟢 Yes' if q_g <= Q_in_max else '🔴 No',
                        })
            if rows_s4:
                st.table(rows_s4)

            # Power–Pressure graph
            st.subheader("Power vs Pressure — Gas Comparison")
            P_vals_s4 = np.linspace(1.0, P_max_s4, 25)
            colors_gas = {'Air': '#1565C0', 'Helium': '#C62828', 'Hydrogen': '#2E7D32'}
            lf_pp = dict(flow=True, regen_imp=True, mechanical=True,
                         wall_cond=True, leakage=True, shuttle=False)

            fig_pp, ax_pp = plt.subplots(figsize=(9, 5))
            fig_pp.patch.set_facecolor('white'); ax_pp.set_facecolor('white')
            for gas_name in ['Air', 'Helium', 'Hydrogen']:
                r = s4.get(gas_name, {})
                bp_g = r.get('best_params')
                if bp_g is None:
                    continue
                sw_pp = pressure_sweep(
                    {**proto2_params, **{k: bp_g[k] for k in ('T_h', 'f')
                                         if k in bp_g}, 'gas': gas_name},
                    P_vals_s4, 'schmidt', lf_pp, Q_in_max)
                if len(sw_pp.get('P_mean_bar', [])) > 0:
                    ax_pp.plot(sw_pp['P_mean_bar'], sw_pp['P_brake'],
                               color=colors_gas[gas_name], lw=2, label=gas_name)
                    infeas = ~sw_pp['feasible']
                    if infeas.any():
                        ax_pp.scatter(sw_pp['P_mean_bar'][infeas],
                                      sw_pp['P_brake'][infeas],
                                      color=colors_gas[gas_name], marker='x', s=40)

            ax_pp.set(xlabel='Mean Pressure [bar]', ylabel='P_brake [W]',
                      title='Brake Power vs Pressure (Prototype 2 geometry)')
            ax_pp.legend(); ax_pp.grid(alpha=0.3)
            st.pyplot(fig_pp); plt.close(fig_pp)
            st.caption(
                "× markers = Q_in > budget (infeasible). "
                "Hydrogen: high performance requires strict safety measures."
            )

st.caption("💡 Stirling Engine Simulator v10 — physics_v10.py · optimization_v10.py · "
           "app_v10.py · animation_v10.py")
