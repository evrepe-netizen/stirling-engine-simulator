"""
app_v9_4.py — Stirling Engine Simulator v9.4 (Streamlit UI)
============================================================
Tabs: Schmidt | Adiabatic | Both | Optimization

v9.4 changes:
  - Default heat input raised from 500 W → 1000 W (prevents red deficit on load)
  - Detailed results shown as a clean st.table (not collapsed in expander)
  - Tab layout reordered: Metrics → Detailed Table → Diagrams → Heat Budget → Animation
  - No-Regenerator P-V loop overlaid on Schmidt and Both P-V graphs
"""

import math, io, warnings
from datetime import datetime

import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import streamlit as st

from animation_v9_4 import build_engine_animation
from physics_v9_4 import (
    PROTOTYPE, GASES,
    to_si, build_geometry,
    simulate, simulate_fixed_heat,
    validate_mass_conservation, validate_first_law,
    validate_carnot, validate_pressure_scaling,
)
from optimization_v9_4 import (
    OPTIMIZABLE_PARAMS, geometry_sensitivity,
    coarse_fine_search, lhs_search, bayesian_search,
)


def _freeze_dict(d):
    """Convert a dict to a sorted tuple-of-pairs for use as a cache key."""
    return tuple(sorted((k, v) for k, v in d.items() if isinstance(v, (int, float, str, bool))))

warnings.filterwarnings("ignore", category=UserWarning)

st.set_page_config(
    page_title="Stirling Engine Simulator v9.4",
    layout="wide",
    initial_sidebar_state="expanded",
)

if 'initialized' not in st.session_state:
    for k, v in PROTOTYPE.items():
        st.session_state[k] = v
    st.session_state['gap_mm']                    = PROTOTYPE['gap'] * 1e3
    st.session_state['L_displacer_effective_mm']  = PROTOTYPE['L_displacer_effective'] * 1e3
    st.session_state.initialized = True

def reset_to_prototype():
    for k, v in PROTOTYPE.items():
        st.session_state[k] = v
    st.session_state['gap_mm']                   = PROTOTYPE['gap'] * 1e3
    st.session_state['L_displacer_effective_mm'] = PROTOTYPE['L_displacer_effective'] * 1e3

col_title, col_reset = st.columns([4, 1])
with col_title:
    st.title("🔥 Stirling Engine Simulator v9.4")
    st.caption("Gamma-type engine — Schmidt · Adiabatic · Comparison · Optimization · Bi-directional Mode")
with col_reset:
    st.write(""); st.write("")
    if st.button("🔄 Reset to Prototype", type="secondary",
                 help="Restore all parameters to the measured prototype values."):
        reset_to_prototype(); st.rerun()


# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.header("⚙️ Parameters")

    st.subheader("Geometry")
    st.number_input("Displacer diameter [mm]", 30.0, 200.0, step=1.0, key='D_displacer',
        help="Bore diameter of the displacer (hot) cylinder. Larger bore increases swept volume "
             "and hot-space area, raising power but also shuttle and conduction losses.")
    st.number_input("Displacer stroke [mm]", 20.0, 300.0, step=1.0, key='S_displacer',
        help="Axial stroke of the displacer piston. Directly sets the expansion swept volume V_swe. "
             "Increasing stroke raises work per cycle but also shuttle loss (∝ stroke²).")
    st.number_input("Power piston diameter [mm]", 30.0, 200.0, step=1.0, key='D_power',
        help="Bore of the power (compression) cylinder. Sets the compression swept volume V_swc.")
    st.number_input("Power piston stroke [mm]", 20.0, 300.0, step=1.0, key='S_power',
        help="Stroke of the power piston. Together with its bore, determines V_swc.")
    st.number_input("Phase angle [°]", 45.0, 135.0, step=5.0, key='phi_deg',
        help="Crank phase angle between displacer and power piston (90° is classical). "
             "Affects timing of pressure peaks relative to volume changes.")

    st.subheader("Regenerator")
    st.number_input("Regen diameter [mm]", 10.0, 150.0, step=1.0, key='D_r',
        help="Inner diameter of the regenerator matrix. Sets the flow cross-section A_r. "
             "Larger diameter reduces flow velocity and Ergun pressure drop.")
    st.number_input("Regen length [mm]", 20.0, 500.0, step=5.0, key='L_r',
        help="Axial length of the regenerator. Longer regen improves effectiveness but increases "
             "flow resistance and dead volume. Set < 5 mm to model 'no regenerator'.")
    st.number_input("Wire diameter [mm]", 0.05, 5.0, step=0.05, key='d_wire',
        help="Diameter of the wire mesh or packing material. Finer wire → better heat transfer "
             "but higher Ergun flow resistance.")
    st.slider("Porosity", 0.5, 0.99, step=0.01, key='porosity',
        help="Void fraction of the regenerator matrix (0.9 = 90% open space). "
             "Higher porosity reduces flow resistance but also reduces heat-transfer surface area.")

    st.subheader("Operating Conditions")
    st.selectbox("Working gas", ["Air", "Helium", "Hydrogen"], key='gas')

    driving_mode = st.radio("Driving Mode",
        ["Fixed Hot Temperature (T_h)", "Fixed Heat Input (Q_in)"],
        key='driving_mode')

    if driving_mode == "Fixed Hot Temperature (T_h)":
        st.number_input("Hot temperature T_h [K]", 400, 1500, step=10, key='T_h')
        st.slider("Max heat input budget [W]", 50, 3000, value=1000, step=10, key='Q_in_max',
            help="Heat constraint. If required heat exceeds this, available power is scaled down.")
    else:
        st.slider("Net Heat Transferred to Gas (Q_in) [W]", 50, 3000, value=1000, step=10, key='Q_in_max',
            help="Assumes ideal reservoir immersion. The engine will calculate the resulting T_h.")

    st.number_input("Cold temperature T_k [K]", 250, 400, step=5, key='T_k')
    st.number_input("Mean pressure [bar]", 0.5, 50.0, step=0.5, key='P_mean_bar')
    st.number_input("Frequency [Hz]", 1, 100, step=1, key='f')

    st.subheader("Losses")
    flow_loss  = st.checkbox("Regen flow loss (Ergun)", True,
        help="Viscous and inertial pressure drop through the regenerator packing (Ergun equation). "
             "Subtracts from W_cycle before any other loss.")
    regen_imp  = st.checkbox("Regen imperfection", True,
        help="Extra heat that must be supplied because the regenerator does not perfectly "
             "pre-heat/pre-cool the gas. Q_miss = M·Cv·ΔT·(1−ε_reg). "
             "Labelled 'Estimated Upper-Bound' — uses total gas mass, which overestimates loss.")
    mech_loss  = st.checkbox("Mech friction", True,
        help="Bearing and seal friction. Modelled as a fixed fraction (1−η_mech) of shaft work.")
    wall_cond  = st.checkbox("Wall conduction", True,
        help="Steady-state axial conduction through the regenerator housing wall, "
             "driven by the T_h − T_k temperature gradient.")
    leak_loss  = st.checkbox("Seal leakage", True,
        help="Work lost to gas escaping past piston seals. Proportional to mean pressure "
             "and the empirical leakage coefficient C_leak.")
    shuttle_loss = st.checkbox("Shuttle heat loss", True,
        help="Cyclic heat conduction through the gas film in the displacer–cylinder annular gap "
             "(Organ 1992). Depends on stroke², gap, and gas conductivity.")
    st.slider("ε_reg",  0.5, 0.99, step=0.01, key='eps_reg',
        help="Regenerator thermal effectiveness (0.85 = 85% of ideal heat recovery). "
             "Has no effect when L_r < 5 mm (no-regenerator mode).")
    st.slider("η_mech", 0.5, 0.99, step=0.01, key='eta_mech',
        help="Mechanical efficiency of the crankshaft, bearings, and seals combined.")
    st.slider("C_leak", 0.0, 0.20, step=0.01, key='C_leak',
        help="Seal leakage coefficient. Fraction of net work lost per bar of mean pressure above P_ref.")

    st.subheader("Shuttle Loss Geometry")
    st.number_input("Displacer radial gap [mm]", 0.10, 0.50, step=0.05,
        key='gap_mm',
        help="Radial clearance between the displacer body and cylinder bore [mm]. "
             "Small gap increases friction risk and wear. "
             "Large gap reduces shuttle loss but increases bypass flow and leakage around the displacer. "
             "Bounds: 0.1–0.5 mm enforced to prevent unphysical optimizer solutions.")
    st.number_input("Effective displacer length [mm]", 50.0, 235.0, step=5.0,
        key='L_displacer_effective_mm',
        help="Length of the shuttle heat-loss zone — typically the hexagonal rod section "
             "bounded by graphite seals on the prototype (default 189 mm).")


# ── Build shared params ───────────────────────────────────────────────────────
params = {k: st.session_state[k] for k in PROTOTYPE if k in st.session_state}
if 'T_h' not in params:
    params['T_h'] = PROTOTYPE['T_h']
for k in ('V_loop_cold', 'V_loop_hot', 'V_cle', 'V_clc',
          'L_displacer', 'P_ref', 'k_metal', 't_wall',
          'D_displacer_effective'):
    params[k] = PROTOTYPE[k]

params['gap']                   = st.session_state.get('gap_mm',
                                    PROTOTYPE['gap'] * 1e3) * 1e-3
params['L_displacer_effective'] = st.session_state.get('L_displacer_effective_mm',
                                    PROTOTYPE['L_displacer_effective'] * 1e3) * 1e-3

losses_flags = dict(flow=flow_loss, regen_imp=regen_imp, mechanical=mech_loss,
                    wall_cond=wall_cond, leakage=leak_loss, shuttle=shuttle_loss)

# ── Driving-mode state ────────────────────────────────────────────────────────
fixed_Qin_mode = (st.session_state.get('driving_mode', 'Fixed Hot Temperature (T_h)')
                  == "Fixed Heat Input (Q_in)")
Q_in_max = st.session_state.get('Q_in_max', 1000)

# ── Run simulations ───────────────────────────────────────────────────────────
with st.spinner("Computing..."):
    if fixed_Qin_mode:
        sim_s = simulate_fixed_heat(params, Q_in_max,
                                    model='schmidt', losses_flags=losses_flags)
        params_noreg = dict(params); params_noreg['L_r'] = 0.001
        sim_s_noreg  = simulate_fixed_heat(params_noreg, Q_in_max,
                                           model='schmidt', losses_flags=losses_flags)
        sim_a        = simulate_fixed_heat(params, Q_in_max,
                                           model='adiabatic', losses_flags=losses_flags)
    else:
        sim_s = simulate(params, model='schmidt', losses_flags=losses_flags)
        params_noreg = dict(params); params_noreg['L_r'] = 0.001
        sim_s_noreg  = simulate(params_noreg, model='schmidt', losses_flags=losses_flags)
        sim_a        = simulate(params, model='adiabatic', losses_flags=losses_flags)


# ── Dead-volume total for P-V X-axis ─────────────────────────────────────────
def _dead_vol(sim):
    """Return total dead volume V_k + V_r + V_h [cm³] for true P-V plot."""
    g = sim['geom']
    return (g.get('V_k', 0) + g.get('V_r', 0) + g.get('V_h', 0)) * 1e6


# ── Heat budget helper ────────────────────────────────────────────────────────
def show_heat_budget(losses, Q_in_max, fixed_Qin_mode=False):
    Q_req    = losses['Q_in_W']
    P_brake  = losses['P_brake']
    margin   = Q_in_max - Q_req
    feasible = margin >= -0.1

    if fixed_Qin_mode:
        st.success(f"✅ Heat target matched perfectly. Engine consumes {Q_req:.1f} W.")
        P_available = P_brake
    else:
        P_available = P_brake * min(1.0, Q_in_max / Q_req) if Q_req > 0 else 0.0
        if feasible:
            st.success(f"✅ Heat source is sufficient. Surplus: {margin:.1f} W")
        else:
            st.error(f"❌ Heat source is insufficient. Deficit: {abs(margin):.1f} W")

    st.table({
        "Parameter": [
            "Required heat input (Estimated Upper-Bound) [W]",
            "Available / Target heat input [W]",
            "Heat surplus / deficit [W]",
            "Feasible?",
            "Brake power [W]",
        ],
        "Value": [
            f"{Q_req:.1f}",
            f"{Q_in_max:.1f}",
            f"{margin:+.1f}",
            "Yes" if feasible else "No",
            f"{P_brake:.2f}",
        ]
    })
    st.caption(
        "⚠️ Heat input is an estimated upper bound (Q_miss uses total gas mass). "
        "Real heater power will also include external heat-transfer losses."
    )
    if not fixed_Qin_mode:
        st.info(
            "📌 **Note:** The Heat Budget tool acts only as a constraint check. "
            "In physical reality, if you have a large heat surplus, the engine's "
            "hot temperature (T_h) will rise, which increases power. "
            "To simulate this real-world behaviour, manually increase the **T_h slider**."
        )


# ── Excel export helper ───────────────────────────────────────────────────────
def build_excel(sim_list, params, losses_flags, Q_in_max, include_charts=True):
    """Returns BytesIO of an xlsx workbook."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.drawing.image import Image as XLImage

    wb = openpyxl.Workbook()
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    header_fill = PatternFill("solid", fgColor="1565C0")
    header_font = Font(color="FFFFFF", bold=True)

    def _hdr(ws, row, col, text):
        c = ws.cell(row=row, column=col, value=text)
        c.fill = header_fill; c.font = header_font
        c.alignment = Alignment(horizontal='center')

    def _row(ws, r, label, value, unit=""):
        ws.cell(r, 1, label); ws.cell(r, 2, value); ws.cell(r, 3, unit)

    for sheet_idx, (sim, model_label) in enumerate(sim_list):
        if sim is None:
            continue
        L = sim['losses']

        ws_name = f"Summary_{model_label}" if len(sim_list) > 1 else "Summary"
        ws1 = wb.active if sheet_idx == 0 else wb.create_sheet(ws_name)
        if sheet_idx == 0:
            ws1.title = ws_name

        _hdr(ws1, 1, 1, f"Stirling Engine Simulator v9.4 — {model_label}")
        ws1.merge_cells('A1:C1')
        _row(ws1, 2, "Timestamp", ts)
        _row(ws1, 3, "Model", model_label)
        _row(ws1, 4, "Gas", params['gas'])
        ws1.append([])
        _hdr(ws1, 6, 1, "INPUT PARAMETERS")
        input_rows = [
            ("D_displacer",  params['D_displacer'],  "mm"),
            ("S_displacer",  params['S_displacer'],  "mm"),
            ("D_power",      params['D_power'],      "mm"),
            ("S_power",      params['S_power'],      "mm"),
            ("phi_deg",      params['phi_deg'],      "°"),
            ("D_r",          params['D_r'],          "mm"),
            ("L_r",          params['L_r'],          "mm"),
            ("d_wire",       params['d_wire'],       "mm"),
            ("porosity",     params['porosity'],     ""),
            ("T_h",          params['T_h'],          "K"),
            ("T_k",          params['T_k'],          "K"),
            ("P_mean_bar",   params['P_mean_bar'],   "bar"),
            ("f",            params['f'],            "Hz"),
            ("eps_reg",      params['eps_reg'],      ""),
            ("eta_mech",     params['eta_mech'],     ""),
            ("gap",          params.get('gap', PROTOTYPE['gap']) * 1e3, "mm"),
            ("L_disp_eff",   params.get('L_displacer_effective',
                              PROTOTYPE['L_displacer_effective']) * 1e3, "mm"),
        ]
        for i, (label, val, unit) in enumerate(input_rows, 7):
            _row(ws1, i, label, val, unit)
        r = 7 + len(input_rows) + 1
        _hdr(ws1, r, 1, "KEY OUTPUTS"); r += 1
        for label, val, unit in [
            ("W_cycle",     L['W_cycle'],          "J"),
            ("W_shaft",     L['W_shaft'],          "J"),
            ("P_brake",     L['P_brake'],          "W"),
            ("eta_brake",   L['eta_brake'] * 100,  "%"),
            ("eta_carnot",  L['eta_carnot'] * 100, "%"),
            ("frac_carnot", L['frac_carnot'] * 100,"%"),
            ("M_gas",       L['M'] * 1000,         "g"),
            ("P_mean_out",  L['P_mean'] / 1e5,     "bar"),
        ]:
            _row(ws1, r, label, round(val, 5), unit); r += 1

        _hdr(ws1, r, 1, "HEAT BUDGET (Estimated Upper-Bound)"); r += 1
        Q_req  = L['Q_in_W']
        margin = Q_in_max - Q_req
        P_avail = L['P_brake'] * (Q_in_max / Q_req) if Q_req > 0 else 0.0
        for label, val, unit in [
            ("Required heat input (upper-bound)", Q_req,      "W"),
            ("Available heat input",              Q_in_max,   "W"),
            ("Heat surplus / deficit",            margin,     "W"),
            ("Feasible?",  "Yes" if Q_req <= Q_in_max else "No", ""),
            ("Brake power (no heat limit)",       L['P_brake'], "W"),
            ("Est. brake power (heat limited)",   P_avail,      "W"),
        ]:
            _row(ws1, r, label, val, unit); r += 1

        ws1.column_dimensions['A'].width = 32
        ws1.column_dimensions['B'].width = 18
        ws1.column_dimensions['C'].width = 8

        ws2 = wb.create_sheet(f"Losses_{model_label}")
        _hdr(ws2, 1, 1, "Loss Component")
        _hdr(ws2, 1, 2, "J/cycle")
        _hdr(ws2, 1, 3, "W")
        loss_rows = [
            ("W_pump (flow)",       L['W_pump'],      L['W_pump'] * params['f']),
            ("W_leak (leakage)",    L['W_leak'],      L['W_leak'] * params['f']),
            ("W_mech_loss",         L['W_mech_loss'], L['W_mech_loss'] * params['f']),
            ("Q_miss (regen) [UB]", L['Q_miss'],      L['Q_miss'] * params['f']),
            ("Q_cond (wall)",       L['Q_cond'],      L.get('Q_cond_W', 0)),
            ("Q_shuttle",           L['Q_shuttle'],   L.get('Q_shuttle_W', 0)),
            ("W_shaft (output)",    L['W_shaft'],     L['P_brake']),
        ]
        for i, (label, jc, w) in enumerate(loss_rows, 2):
            ws2.cell(i, 1, label); ws2.cell(i, 2, round(jc, 5)); ws2.cell(i, 3, round(w, 3))
        ws2.column_dimensions['A'].width = 28
        ws2.column_dimensions['B'].width = 14

        ws3 = wb.create_sheet(f"CycleData_{model_label}")
        headers3 = ["theta_deg", "P_bar", "V_e_cm3", "V_c_cm3",
                    "T_e_K", "T_c_K", "V_total_abs_cm3"]
        for j, h in enumerate(headers3, 1):
            _hdr(ws3, 1, j, h)
        R    = sim['result']
        dv   = _dead_vol(sim)
        for i in range(len(R['theta'])):
            ws3.append([
                round(math.degrees(R['theta'][i]), 2),
                round(R['P'][i] / 1e5, 5),
                round(R['V_e'][i] * 1e6, 4),
                round(R['V_c'][i] * 1e6, 4),
                round(R['T_e'][i], 2),
                round(R['T_c'][i], 2),
                round((R['V_e'][i] + R['V_c'][i]) * 1e6 + dv, 4),
            ])

        if include_charts:
            ws4  = wb.create_sheet(f"Charts_{model_label}")
            dv   = _dead_vol(sim)
            V_tot = (R['V_e'] + R['V_c']) * 1e6 + dv
            th_deg = np.rad2deg(R['theta'])

            fig_pv, ax_pv = plt.subplots(figsize=(8, 5))
            ax_pv.plot(V_tot, R['P'] / 1e5, '#1565C0', lw=2)
            ax_pv.fill(V_tot, R['P'] / 1e5, alpha=0.1, color='#1565C0')
            ax_pv.set(xlabel='V_total (absolute) [cm³]', ylabel='P [bar]',
                      title=f'P-V Diagram — {model_label}')
            ax_pv.grid(alpha=0.3)
            buf_pv = io.BytesIO()
            fig_pv.savefig(buf_pv, format='png', dpi=100, bbox_inches='tight')
            plt.close(fig_pv); buf_pv.seek(0)
            ws4.add_image(XLImage(buf_pv), 'A1')

            fig_pth, ax_pth = plt.subplots(figsize=(8, 5))
            ax_pth.plot(th_deg, R['P'] / 1e5, '#1565C0', lw=2)
            ax_pth.set(xlabel='θ [°]', ylabel='P [bar]',
                       title=f'P vs θ — {model_label}')
            ax_pth.grid(alpha=0.3)
            buf_pth = io.BytesIO()
            fig_pth.savefig(buf_pth, format='png', dpi=100, bbox_inches='tight')
            plt.close(fig_pth); buf_pth.seek(0)
            ws4.add_image(XLImage(buf_pth), 'A32')

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf


# ── Tabs ──────────────────────────────────────────────────────────────────────
tab_schmidt, tab_adiabatic, tab_both, tab_optimize = st.tabs([
    "📊 Schmidt (Isothermal)",
    "🔁 Adiabatic (RK45)",
    "⚖️ Both Models",
    "🎯 Optimization",
])


# ════════════════════════════════════════════════════════════════════════════
# TAB 1 — SCHMIDT
# ════════════════════════════════════════════════════════════════════════════
with tab_schmidt:
    st.header("📊 Schmidt (Isothermal) Model")

    use_regen = st.checkbox("With regenerator / heat recovery", value=True, key='schmidt_regen',
        help="When unchecked, sets L_r = 1 mm (< 5 mm threshold) which forces eps_reg = 0 "
             "AND Q_cond = 0 (no regenerator housing exists). "
             "This models the thermodynamic effect of removing heat recovery entirely.")
    sim = sim_s if use_regen else sim_s_noreg

    if sim is None:
        if fixed_Qin_mode:
            st.error(
                "❌ Root-finding failed. The target heat input may be outside the physically "
                f"achievable range for this engine configuration. "
                f"Try a Q_in target between ~10 W and ~500 W, or check that T_k < 400 K "
                f"and mean pressure ≥ 0.5 bar."
            )
        else:
            st.error("Simulation failed — check parameters.")
    else:
        L = sim['losses']

        # ── 1. Mode banner ────────────────────────────────────────────────────
        if fixed_Qin_mode:
            T_h_solved = sim.get('T_h_solved', sim['gas']['T_h'])
            st.success(
                f"✅ **Fixed Heat Input mode** — target Q_in = **{Q_in_max} W**  →  "
                f"Solved T_h = **{T_h_solved:.1f} K**  "
                f"(converged in {sim.get('solver_iters', '?')} iterations)"
            )

        # ── 2. Top Metrics ────────────────────────────────────────────────────
        c0, c1, c2, c3, c4 = st.columns(5)
        c0.metric("W_cycle [J]", f"{L['W_cycle']:.4f}")
        c1.metric("W_shaft [J]", f"{L['W_shaft']:.4f}")
        c2.metric("Brake Power [W]", f"{L['P_brake']:.2f}")
        c3.metric("η_brake [%]", f"{L['eta_brake']*100:.2f}")

        fixed_Qin_mode = (st.session_state.get('driving_mode') == "Fixed Heat Input (Q_in)")
        if fixed_Qin_mode and 'T_e_max' in L:
            c4.metric("🔥 Calculated T_h [K]", f"{sim['gas']['T_h']:.1f}")

        # ── 3. Detailed Results Table ─────────────────────────────────────────
        st.subheader("📋 Detailed Results")
        st.table({
            "Parameter": [
                "Mass (gas mass)",
                "Mean Pressure",
                "W_cycle",
                "W_shaft",
                "Brake Power",
                "Q_in [Upper-Bound]",
                "Q_miss [Upper-Bound]",
                "Q_shuttle",
                "Efficiency",
                "% Carnot",
            ],
            "Value": [
                f"{L['M']*1000:.4f} g",
                f"{L['P_mean']/1e5:.3f} bar",
                f"{L['W_cycle']:.4f} J",
                f"{L['W_shaft']:.4f} J",
                f"{L['P_brake']:.2f} W",
                f"{L['Q_in_W']:.2f} W (est. upper-bound)",
                f"{L['Q_miss']*params['f']:.2f} W (est. upper-bound)",
                f"{L.get('Q_shuttle_W', 0):.2f} W",
                f"{L['eta_brake']*100:.3f} %",
                f"{L['frac_carnot']*100:.2f} %",
            ],
        })

        # ── 4. Diagrams ───────────────────────────────────────────────────────
        st.subheader("📈 Diagrams")
        R    = sim['result']
        dv   = _dead_vol(sim)
        V_tot = (R['V_e'] + R['V_c']) * 1e6 + dv
        th_deg = np.rad2deg(R['theta'])

        fig, axes = plt.subplots(1, 3, figsize=(15, 5))
        fig.patch.set_facecolor('white')

        ax = axes[0]
        ax.set_facecolor('white')
        ax.plot(V_tot, R['P'] / 1e5, '#1565C0', lw=2, label='Schmidt')
        ax.fill(V_tot, R['P'] / 1e5, alpha=0.12, color='#1565C0')
        # No-Regen overlay
        if sim_s_noreg is not None:
            R_nr = sim_s_noreg['result']
            dv_nr = _dead_vol(sim_s_noreg)
            V_tot_nr = (R_nr['V_e'] + R_nr['V_c']) * 1e6 + dv_nr
            ax.plot(V_tot_nr, R_nr['P'] / 1e5, color='#FFA726', lw=2, ls='--', label='No Regen')
            ax.legend(fontsize=8)
        ax.set(xlabel='V_total — absolute [cm³]', ylabel='P [bar]', title='P-V Diagram')
        ax.grid(alpha=0.3)

        ax = axes[1]
        ax.set_facecolor('white')
        ax.plot(th_deg, R['P'] / 1e5, '#1565C0', lw=2)
        ax.set(xlabel='θ [°]', ylabel='P [bar]', title='Pressure vs θ')
        ax.grid(alpha=0.3)

        ax = axes[2]
        ax.set_facecolor('white')
        Q_e_val       = abs(L.get('Q_e', 0) * params['f'])
        Q_miss_val    = abs(L.get('Q_miss', 0) * params['f'])
        Q_cond_val    = abs(L.get('Q_cond_W', 0))
        Q_shuttle_val = abs(L.get('Q_shuttle_W', 0))
        labels = ['Q_e (expansion) [UB]', 'Q_miss (regen) [UB]', 'Q_cond (wall)', 'Q_shuttle']
        values = [Q_e_val, Q_miss_val, Q_cond_val, Q_shuttle_val]
        colors = ['#EF5350', '#FFA726', '#42A5F5', '#AB47BC']
        ax.bar(labels, values, color=colors, edgecolor='#333', linewidth=0.8)
        ax.set(ylabel='Power [W]', title='Heat Components')
        ax.tick_params(axis='x', labelrotation=15)
        ax.grid(alpha=0.3, axis='y')

        plt.tight_layout()
        st.pyplot(fig); plt.close(fig)

        # ── 5. Heat Input Budget ──────────────────────────────────────────────
        st.subheader("🔥 Heat Input Budget — Estimated Upper-Bound (Schmidt)")
        show_heat_budget(L, Q_in_max, fixed_Qin_mode=fixed_Qin_mode)

        # ── 6. Engine Animation ───────────────────────────────────────────────
        st.subheader("🎬 Engine Animation")
        if st.button("Generate Animation"):
            with st.spinner("Rendering animation..."):
                try:
                    geom_frozen   = _freeze_dict(build_geometry(to_si(params)))
                    params_frozen = _freeze_dict(params)
                    gif_b64 = build_engine_animation(geom_frozen, params_frozen)
                    st.markdown(
                        f'<img src="data:image/gif;base64,{gif_b64}" '
                        f'style="width:100%;max-width:780px;border-radius:8px;" />',
                        unsafe_allow_html=True,
                    )
                except Exception as e:
                    st.warning(f"Animation error: {e}")

        # ── 7. Top-3 Geometric Improvements ──────────────────────────────────
        st.subheader("🔧 Top-3 Geometric Improvements (Global Sweep)")
        st.caption("Full-range sweep per geometric parameter — operating conditions held fixed.")
        with st.spinner("Running global sensitivity analysis..."):
            sens = geometry_sensitivity(params, losses_flags)
        if sens:
            for rank, (key, name, units, base_val, best_val, delta_W, pct) in enumerate(sens[:3], 1):
                direction = "↑ Increase" if best_val > base_val else "↓ Decrease"
                arrow = "🟢" if delta_W > 0 else "🔴"
                with st.expander(
                    f"#{rank}  {name}  —  {arrow} {delta_W:+.2f} W  ({pct:+.1f}%)",
                    expanded=(rank == 1)
                ):
                    st.markdown(
                        f"**{direction} {name}** from **{base_val:.2f} {units}** "
                        f"to **{best_val:.2f} {units}**\n\n"
                        f"Estimated brake power change: **{delta_W:+.2f} W** ({pct:+.1f}%)\n\n"
                        f"*Based on full-range sweep (Schmidt model, all other params fixed).*"
                    )
        else:
            st.info("Could not compute sensitivity — check parameters.")

        # ── 8. Power vs Mean Pressure sensitivity ─────────────────────────────
        st.subheader("📉 Power vs Mean Pressure (All Gases)")
        st.caption("Quick parametric sweep — all other parameters held constant at current values.")
        with st.spinner("Running pressure sweep..."):
            lf_sweep = dict(flow=flow_loss, regen_imp=regen_imp, mechanical=mech_loss,
                            wall_cond=wall_cond, leakage=leak_loss, shuttle=shuttle_loss)
            P_range = np.linspace(0.5, min(30.0, params['P_mean_bar'] * 4), 20)
            gas_colors = {'Air': '#1565C0', 'Helium': '#C62828', 'Hydrogen': '#2E7D32'}

            fig_ps, ax_ps = plt.subplots(figsize=(9, 5))
            fig_ps.patch.set_facecolor('white'); ax_ps.set_facecolor('white')
            for gas_name, gcolor in gas_colors.items():
                pw_vals = []
                for pm in P_range:
                    p_sw = dict(params); p_sw['gas'] = gas_name; p_sw['P_mean_bar'] = float(pm)
                    s_sw = simulate(p_sw, model='schmidt', losses_flags=lf_sweep)
                    pw_vals.append(s_sw['losses']['P_brake'] if s_sw else float('nan'))
                ax_ps.plot(P_range, pw_vals, color=gcolor, lw=2, label=gas_name)

            ax_ps.axvline(params['P_mean_bar'], color='#888', ls='--', lw=1.2,
                          label=f"Current P_mean = {params['P_mean_bar']:.1f} bar")
            ax_ps.set(xlabel='Mean Pressure [bar]', ylabel='Brake Power [W]',
                      title='Brake Power vs Mean Pressure — Schmidt model')
            ax_ps.legend(fontsize=9); ax_ps.grid(alpha=0.3)
            plt.tight_layout()
            st.pyplot(fig_ps); plt.close(fig_ps)

        if st.button("📥 Export to Excel (Schmidt)"):
            try:
                buf = build_excel([(sim, "Schmidt")], params, losses_flags, Q_in_max)
                fname = f"stirling_v9_4_schmidt_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                st.download_button("⬇️ Download Excel (Schmidt)", data=buf.getvalue(),
                    file_name=fname,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            except Exception as e:
                st.error(f"Export error: {e}")


# ════════════════════════════════════════════════════════════════════════════
# TAB 2 — ADIABATIC
# ════════════════════════════════════════════════════════════════════════════
with tab_adiabatic:
    st.header("🔁 Adiabatic (RK45) Model")

    if sim_a is None:
        if fixed_Qin_mode:
            st.error(
                "❌ Adiabatic root-finding failed. The heat target may be unreachable "
                "with the current geometry and pressure. Try the Schmidt model first to "
                "verify feasibility, then switch to Adiabatic."
            )
        else:
            st.error("Adiabatic simulation failed — check parameters.")
    else:
        La = sim_a['losses']
        Ls = sim_s['losses'] if sim_s else None

        # ── 1. Mode banner ────────────────────────────────────────────────────
        if fixed_Qin_mode:
            T_h_solved_a = sim_a.get('T_h_solved', sim_a['gas']['T_h'])
            st.success(
                f"✅ **Fixed Heat Input mode** — target Q_in = **{Q_in_max} W**  →  "
                f"Solved T_h = **{T_h_solved_a:.1f} K**  "
                f"(converged in {sim_a.get('solver_iters', '?')} iterations)"
            )

        # ── 2. Top Metrics ────────────────────────────────────────────────────
        c0, c1, c2, c3, c4 = st.columns(5)
        c0.metric("W_cycle [J]", f"{La['W_cycle']:.4f}")
        c1.metric("W_shaft [J]", f"{La['W_shaft']:.4f}")
        c2.metric("Brake Power [W]", f"{La['P_brake']:.2f}")
        c3.metric("η_brake [%]", f"{La['eta_brake']*100:.2f}")

        fixed_Qin_mode = (st.session_state.get('driving_mode') == "Fixed Heat Input (Q_in)")
        if fixed_Qin_mode and 'T_e_max' in La:
            c4.metric("🔥 Calculated T_h [K]", f"{sim_a['gas']['T_h']:.1f}")

        # ── 3. Detailed Results Table ─────────────────────────────────────────
        st.subheader("📋 Detailed Results")
        st.table({
            "Parameter": [
                "Mass (gas mass)",
                "Mean Pressure",
                "W_cycle",
                "W_shaft",
                "Brake Power",
                "Q_in [Upper-Bound]",
                "Q_miss [Upper-Bound]",
                "Q_shuttle",
                "Efficiency",
                "% Carnot",
            ],
            "Value": [
                f"{La['M']*1000:.4f} g",
                f"{La['P_mean']/1e5:.3f} bar",
                f"{La['W_cycle']:.4f} J",
                f"{La['W_shaft']:.4f} J",
                f"{La['P_brake']:.2f} W",
                f"{La['Q_in_W']:.2f} W (est. upper-bound)",
                f"{La['Q_miss']*params['f']:.2f} W (est. upper-bound)",
                f"{La.get('Q_shuttle_W', 0):.2f} W",
                f"{La['eta_brake']*100:.3f} %",
                f"{La['frac_carnot']*100:.2f} %",
            ],
        })

        if Ls:
            M_g     = Ls['M'] * 1000
            t_ratio = params['T_h'] / params['T_k']
            base_msg = (
                f"⚖️ **Both models share the same gas mass M = {M_g:.4f} g** "
                f"(computed by Schmidt for P_mean = {params['P_mean_bar']:.2f} bar). "
                f"Adiabatic P_mean = {La['P_mean']/1e5:.3f} bar (output, not forced). "
                f"T_h/T_k = {t_ratio:.2f}."
            )
            if La['W_cycle'] > Ls['W_cycle']:
                st.warning(
                    base_msg + f"\n\n"
                    f"⚠️ **W_adiabatic ({La['W_cycle']:.3f} J) > W_schmidt ({Ls['W_cycle']:.3f} J).** "
                    f"Adiabatic work is higher than Schmidt in this run. "
                    f"This may indicate model sensitivity or assumptions in the adiabatic solver "
                    f"(e.g. unconverged initial conditions, extreme T_h/T_k ratio). "
                    f"Validate results before using for design conclusions."
                )
            else:
                st.info(
                    base_msg + f" "
                    f"W_schmidt ({Ls['W_cycle']:.3f} J) ≥ W_adiabatic ({La['W_cycle']:.3f} J)."
                )

        # ── 4. Diagrams ───────────────────────────────────────────────────────
        st.subheader("📈 Diagrams")
        Ra    = sim_a['result']
        dv_a  = _dead_vol(sim_a)
        V_tot_a = (Ra['V_e'] + Ra['V_c']) * 1e6 + dv_a
        th_deg  = np.rad2deg(Ra['theta'])

        fig, axes = plt.subplots(1, 3, figsize=(15, 5))
        fig.patch.set_facecolor('white')

        ax = axes[0]
        ax.set_facecolor('white')
        ax.plot(V_tot_a, Ra['P'] / 1e5, '#C62828', lw=2)
        ax.fill(V_tot_a, Ra['P'] / 1e5, alpha=0.12, color='#C62828')
        ax.set(xlabel='V_total — absolute [cm³]', ylabel='P [bar]', title='P-V Diagram')
        ax.grid(alpha=0.3)

        ax = axes[1]
        ax.set_facecolor('white')
        ax.plot(th_deg, Ra['P'] / 1e5, '#C62828', lw=2)
        ax.set(xlabel='θ [°]', ylabel='P [bar]', title='Pressure vs θ')
        ax.grid(alpha=0.3)

        ax = axes[2]
        ax.set_facecolor('white')
        ax.plot(th_deg, Ra['T_e'], '#D32F2F', lw=2, label='T_e')
        ax.plot(th_deg, Ra['T_c'], '#1976D2', lw=2, label='T_c')
        ax.axhline(params['T_h'], color='#D32F2F', ls='--', lw=1.2, alpha=0.6, label='T_h (ref)')
        ax.axhline(params['T_k'], color='#1976D2', ls='--', lw=1.2, alpha=0.6, label='T_k (ref)')
        ax.set(xlabel='θ [°]', ylabel='T [K]', title='Gas Temperatures vs θ')
        ax.legend(fontsize=8); ax.grid(alpha=0.3)

        plt.tight_layout()
        st.pyplot(fig); plt.close(fig)

        # ── 5. Heat Input Budget ──────────────────────────────────────────────
        st.subheader("🔥 Heat Input Budget — Estimated Upper-Bound (Adiabatic)")
        show_heat_budget(La, Q_in_max, fixed_Qin_mode=fixed_Qin_mode)

        # ── 6. Engine Animation ───────────────────────────────────────────────
        st.subheader("🎬 Engine Animation")

        with st.expander("🔍 Validation checks (click to expand)", expanded=False):
            delta, ok = validate_mass_conservation(sim_a['result'], sim_a['geom'], sim_a['gas'])
            (st.success if ok else st.error)(
                f"{'✅' if ok else '❌'}  Mass conservation: variation = **{delta:.3f} %**"
                + (" < 2 % ✓" if ok else " > 2 % limit")
            )
            err, ok = validate_first_law(La)
            (st.success if ok else st.error)(
                f"{'✅' if ok else '❌'}  First-law error: **{err:.6f} %**"
            )
            eta_b, eta_c, ok = validate_carnot(La)
            (st.success if ok else st.error)(
                f"{'✅' if ok else '❌'}  η_brake = {eta_b*100:.3f} %  "
                f"{'≤' if ok else '>'} η_Carnot = {eta_c*100:.2f} %  "
                f"({La['frac_carnot']*100:.1f} % of Carnot)"
            )
            lf0 = dict(flow=False, regen_imp=False, mechanical=False,
                       wall_cond=False, leakage=False, shuttle=False)
            pz  = dict(params); pz['T_h'] = params['T_k'] + 5
            sz  = simulate(pz, model='adiabatic', losses_flags=lf0)
            if sz:
                ratio_z = abs(sz['losses']['W_cycle']) / max(abs(La['W_cycle']), 1e-9)
                ok_z    = ratio_z < 0.05
                (st.success if ok_z else st.error)(
                    f"{'✅' if ok_z else '❌'}  Zero-ΔT: W(ΔT=5K) = {sz['losses']['W_cycle']:.5f} J "
                    f"({ratio_z*100:.2f} % of nominal)"
                )
            W1, W2, ratio_p, ok_p = validate_pressure_scaling(params, losses_flags)
            if W1:
                (st.success if ok_p else st.error)(
                    f"{'✅' if ok_p else '❌'}  "
                    f"Pressure scaling ratio = {ratio_p:.5f} (expected 2.000)"
                )

        if st.button("📥 Export to Excel (Adiabatic)"):
            try:
                buf   = build_excel([(sim_a, "Adiabatic")], params, losses_flags, Q_in_max)
                fname = f"stirling_v9_4_adiabatic_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                st.download_button("⬇️ Download Excel (Adiabatic)", data=buf.getvalue(),
                    file_name=fname,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            except Exception as e:
                st.error(f"Export error: {e}")


# ════════════════════════════════════════════════════════════════════════════
# TAB 3 — BOTH MODELS
# ════════════════════════════════════════════════════════════════════════════
with tab_both:
    st.header("⚖️ Both Models — Side-by-Side Comparison")

    use_regen_both = st.checkbox(
        "Schmidt: with regenerator / heat recovery", value=True, key='both_regen',
        help="Unchecked → no-regen mode: L_r forced < 5 mm, eps_reg = 0, Q_cond = 0.")
    sim_s_display = sim_s if use_regen_both else sim_s_noreg

    if sim_s_display and sim_a:
        Ls = sim_s_display['losses']
        La = sim_a['losses']

        # ── Mode banner ───────────────────────────────────────────────────────
        if fixed_Qin_mode:
            Th_s = sim_s_display.get('T_h_solved', sim_s_display['gas']['T_h'])
            Th_a = sim_a.get('T_h_solved', sim_a['gas']['T_h'])
            st.info(
                f"🔁 **Fixed Heat Input mode** — Q_in target = **{Q_in_max} W**\n\n"
                f"Schmidt solved T_h = **{Th_s:.1f} K** | "
                f"Adiabatic solved T_h = **{Th_a:.1f} K**\n\n"
                "Note: The two models find slightly different T_h values because their "
                "thermodynamic cycles produce different Q_in for the same T_h."
            )

        # ── Comparison table ─────────────────────────────────────────────────
        st.subheader("Comparison Table")
        st.table({
            "Metric":     ["M [g]", "W_cycle [J]", "W_shaft [J]", "P_brake [W]",
                           "η_brake [%]", "P_mean_out [bar]"],
            "Schmidt":    [f"{Ls['M']*1000:.4f}", f"{Ls['W_cycle']:.4f}",
                           f"{Ls['W_shaft']:.4f}", f"{Ls['P_brake']:.2f}",
                           f"{Ls['eta_brake']*100:.2f}", f"{Ls['P_mean']/1e5:.3f}"],
            "Adiabatic":  [f"{La['M']*1000:.4f}", f"{La['W_cycle']:.4f}",
                           f"{La['W_shaft']:.4f}", f"{La['P_brake']:.2f}",
                           f"{La['eta_brake']*100:.2f}", f"{La['P_mean']/1e5:.3f}"],
        })

        t_ratio  = params['T_h'] / params['T_k']
        base_msg = (
            f"⚖️ **Both models share M = {Ls['M']*1000:.4f} g** "
            f"(Schmidt for P_mean = {params['P_mean_bar']:.2f} bar). "
            f"Adiabatic P_mean = {La['P_mean']/1e5:.3f} bar (output). "
            f"T_h/T_k = {t_ratio:.2f}."
        )
        if La['W_cycle'] > Ls['W_cycle']:
            st.warning(
                base_msg + f"\n\n"
                f"⚠️ **W_adiabatic ({La['W_cycle']:.3f} J) > W_schmidt ({Ls['W_cycle']:.3f} J).** "
                f"Adiabatic work is higher than Schmidt in this run. "
                f"This may indicate model sensitivity or assumptions in the adiabatic solver. "
                f"Validate before using for design conclusions."
            )
        else:
            st.info(
                base_msg + f" "
                f"W_schmidt ({Ls['W_cycle']:.3f} J) ≥ W_adiabatic ({La['W_cycle']:.3f} J)."
            )

        # ── Overlaid P-V and Temperature plots ───────────────────────────────
        st.subheader("📈 Overlaid Comparison Plots")

        Rs   = sim_s_display['result']
        Ra   = sim_a['result']
        dv_s = _dead_vol(sim_s_display)
        dv_a = _dead_vol(sim_a)

        V_s  = (Rs['V_e'] + Rs['V_c']) * 1e6 + dv_s
        V_a  = (Ra['V_e'] + Ra['V_c']) * 1e6 + dv_a
        th_s = np.rad2deg(Rs['theta'])
        th_a = np.rad2deg(Ra['theta'])

        fig_both, axes_both = plt.subplots(1, 2, figsize=(14, 5))
        fig_both.patch.set_facecolor('white')

        # P-V overlay — Schmidt + Adiabatic + No-Regen
        ax = axes_both[0]
        ax.set_facecolor('white')
        ax.plot(V_s, Rs['P'] / 1e5, '#1565C0', lw=2, label='Schmidt')
        ax.fill(V_s, Rs['P'] / 1e5, alpha=0.08, color='#1565C0')
        ax.plot(V_a, Ra['P'] / 1e5, '#C62828', lw=2, ls='--', label='Adiabatic')
        ax.fill(V_a, Ra['P'] / 1e5, alpha=0.08, color='#C62828')
        if sim_s_noreg is not None:
            R_nr = sim_s_noreg['result']
            dv_nr = _dead_vol(sim_s_noreg)
            V_tot_nr = (R_nr['V_e'] + R_nr['V_c']) * 1e6 + dv_nr
            ax.plot(V_tot_nr, R_nr['P'] / 1e5, color='#FFA726', lw=2, ls='--', label='No Regen')
        ax.set(xlabel='V_total — absolute [cm³]', ylabel='P [bar]',
               title='P-V Diagram — Overlay')
        ax.legend(fontsize=9); ax.grid(alpha=0.3)

        # Temperature overlay (adiabatic T_e, T_c vs Schmidt flat walls)
        ax = axes_both[1]
        ax.set_facecolor('white')
        ax.plot(th_a, Ra['T_e'], '#D32F2F', lw=2, label='T_e (Adiabatic)')
        ax.plot(th_a, Ra['T_c'], '#1976D2', lw=2, label='T_c (Adiabatic)')
        ax.axhline(params['T_h'], color='#D32F2F', ls='--', lw=1.5, alpha=0.55,
                   label=f"T_h = {params['T_h']} K (Schmidt/wall)")
        ax.axhline(params['T_k'], color='#1976D2', ls='--', lw=1.5, alpha=0.55,
                   label=f"T_k = {params['T_k']} K (Schmidt/wall)")
        ax.set(xlabel='θ [°]', ylabel='T [K]',
               title='Gas Temperatures — Adiabatic vs Schmidt Walls')
        ax.legend(fontsize=8); ax.grid(alpha=0.3)

        plt.tight_layout()
        st.pyplot(fig_both); plt.close(fig_both)

        # ── Heat budgets side by side ─────────────────────────────────────────
        st.subheader("🔥 Heat Input Budget — Estimated Upper-Bound")
        col1, col2 = st.columns(2)
        with col1:
            st.markdown("**Schmidt**")
            show_heat_budget(Ls, Q_in_max, fixed_Qin_mode=fixed_Qin_mode)
        with col2:
            st.markdown("**Adiabatic**")
            show_heat_budget(La, Q_in_max, fixed_Qin_mode=fixed_Qin_mode)

        if st.button("📥 Export to Excel (Both Models)"):
            try:
                buf = build_excel(
                    [(sim_s_display, "Schmidt"), (sim_a, "Adiabatic")],
                    params, losses_flags, Q_in_max
                )
                fname = f"stirling_v9_4_both_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                st.download_button("⬇️ Download Excel (Both Models)", data=buf.getvalue(),
                    file_name=fname,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            except Exception as e:
                st.error(f"Export error: {e}")
    else:
        st.error("One or both simulations failed — check parameters.")


# ════════════════════════════════════════════════════════════════════════════
# TAB 4 — OPTIMIZATION
# ════════════════════════════════════════════════════════════════════════════
with tab_optimize:
    st.header("🎯 Optimization")

    st.subheader("Lock / Open parameters")
    open_specs = []
    n_cols = 2
    cols   = st.columns(n_cols)
    for i, (display_name, key, pmin, pmax, step, units) in enumerate(OPTIMIZABLE_PARAMS):
        col = cols[i % n_cols]
        with col:
            current = st.session_state.get(key, PROTOTYPE.get(key, pmin))
            locked  = st.checkbox(
                f"🔒 Lock **{display_name}** = {current} {units}",
                value=True, key=f"lock_{key}")
            if not locked:
                cc1, cc2 = st.columns(2)
                pmn = cc1.number_input("min", value=float(pmin), step=float(step),
                                       key=f"min_{key}", label_visibility="collapsed")
                pmx = cc2.number_input("max", value=float(pmax), step=float(step),
                                       key=f"max_{key}", label_visibility="collapsed")
                open_specs.append((key, pmn, pmx))

    st.subheader("Strategy")
    strategy = st.radio("Search method:", [
        "Coarse → Fine grid (recommended, fast)",
        "Latin Hypercube Sampling (LHS, smarter)",
        "Bayesian Optimization (advanced, requires scikit-optimize)",
    ], horizontal=False)

    if strategy.startswith("Latin"):
        n_lhs = st.slider("Number of LHS samples", 100, 2000, 500, 100)
    elif strategy.startswith("Bayesian"):
        n_bay = st.slider("Number of Bayesian calls", 20, 200, 60, 10)

    st.subheader("Objective")
    obj_choice = st.radio("Maximize:", [
        "Brake Power", "Brake Efficiency", "Power × Efficiency (balanced)"
    ], horizontal=True)
    obj_map   = {'Brake Power': 'power', 'Brake Efficiency': 'efficiency',
                 'Power × Efficiency (balanced)': 'balanced'}
    objective = obj_map[obj_choice]

    st.subheader("Model")
    opt_model_key = 'schmidt' if st.radio(
        "Model:", ["Schmidt (fast)", "Adiabatic (slower)"], horizontal=True
    ).startswith("Schmidt") else 'adiabatic'

    st.caption(
        "**Excluded from automated optimization:** porosity, ε_reg, η_mech, C_leak — "
        "coupled material/manufacturing properties. Adjust manually in the sidebar.\n\n"
        "**gap bounds: 0.1–0.5 mm** (enforced to prevent unphysical optimizer solutions "
        "that eliminate shuttle loss via an unrealistically large gap)."
    )

    if not open_specs:
        st.warning("Unlock at least one parameter above.")
    else:
        st.info(f"{len(open_specs)} parameter(s) open for optimization.")

    if fixed_Qin_mode:
        st.info(
            "ℹ️ Optimization runs in **Fixed T_h mode** internally "
            f"(T_h = {sim_s['gas']['T_h']:.0f} K — the T_h solved for your Q_in target). "
            "The optimizer maximises brake power at that solved temperature."
        )

    if st.button("🚀 Run Optimization", type="primary", disabled=(len(open_specs) == 0)):
        base_params = dict(params)
        if fixed_Qin_mode and sim_s is not None:
            base_params['T_h'] = sim_s.get('T_h_solved', sim_s['gas']['T_h'])
        opt_flags   = dict(flow=True, regen_imp=True, mechanical=True,
                           wall_cond=True, leakage=True, shuttle=True)
        prog_bar = st.progress(0, text="Starting...")

        def cb(frac, msg):
            prog_bar.progress(min(frac, 1.0), text=msg)

        opt_driving_mode = st.session_state.get('driving_mode', 'Fixed Hot Temperature (T_h)')
        if strategy.startswith("Coarse"):
            bp, bl, all_r = coarse_fine_search(base_params, open_specs, objective,
                                               opt_model_key, opt_flags,
                                               driving_mode=opt_driving_mode,
                                               progress_cb=cb)
        elif strategy.startswith("Latin"):
            bp, bl, all_r = lhs_search(base_params, open_specs, objective,
                                       opt_model_key, opt_flags, n_lhs,
                                       driving_mode=opt_driving_mode,
                                       progress_cb=cb)
        else:
            bp, bl, all_r = bayesian_search(base_params, open_specs, objective,
                                            opt_model_key, opt_flags, n_bay,
                                            driving_mode=opt_driving_mode,
                                            progress_cb=cb)
        prog_bar.empty()

        if bl is None:
            st.error("❌ No valid configuration found.")
        else:
            st.success(f"✅ Done!  Best brake power: {bl['P_brake']:.2f} W, "
                       f"η = {bl['eta_brake']*100:.2f}%")
            sim_cur = simulate(base_params, model=opt_model_key, losses_flags=opt_flags)
            c1, c2  = st.columns(2)
            if sim_cur:
                c1.metric("Best Power", f"{bl['P_brake']:.2f} W",
                          delta=f"{bl['P_brake']-sim_cur['losses']['P_brake']:+.2f} W")
                c2.metric("Best η",     f"{bl['eta_brake']*100:.2f}%",
                          delta=f"{(bl['eta_brake']-sim_cur['losses']['eta_brake'])*100:+.2f}%")

            rows = []
            for _, key, *_ in OPTIMIZABLE_PARAMS:
                if any(s[0] == key for s in open_specs):
                    rows.append({'Parameter': key,
                                 'Current':  f"{base_params.get(key, '?')}",
                                 'Optimum':  f"{bp.get(key, '?')}"})
            st.table(rows)

            sim_best = simulate(bp, model=opt_model_key, losses_flags=opt_flags)
            if sim_cur and sim_best:
                dv_c = _dead_vol(sim_cur)
                dv_b = _dead_vol(sim_best)
                fig2, ax2 = plt.subplots(figsize=(9, 5))
                fig2.patch.set_facecolor('white'); ax2.set_facecolor('white')
                Vc = (sim_cur['result']['V_e']  + sim_cur['result']['V_c'])  * 1e6 + dv_c
                Vb = (sim_best['result']['V_e'] + sim_best['result']['V_c']) * 1e6 + dv_b
                ax2.plot(Vc, sim_cur['result']['P']  / 1e5, '#888',    lw=2,   label='Current')
                ax2.fill(Vc, sim_cur['result']['P']  / 1e5, alpha=0.1, color='#888')
                ax2.plot(Vb, sim_best['result']['P'] / 1e5, '#2E7D32', lw=2.5, label='Optimum')
                ax2.fill(Vb, sim_best['result']['P'] / 1e5, alpha=0.15,color='#2E7D32')
                ax2.set(xlabel='V_total — absolute [cm³]', ylabel='P [bar]',
                        title='P-V: Current vs Optimum')
                ax2.legend(); ax2.grid(alpha=0.3)
                st.pyplot(fig2); plt.close(fig2)

            st.session_state['opt_result'] = dict(
                best_params=bp, best_losses=bl, all_results=all_r,
                base_params=base_params, strategy=strategy, objective=obj_choice
            )

st.caption("💡 Stirling Engine Simulator v9.4 — physics_v9_4.py · optimization_v9_4.py · app_v9_4.py · animation_v9_4.py")
